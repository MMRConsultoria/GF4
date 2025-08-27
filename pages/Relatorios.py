# pages/PainelResultados.py
import streamlit as st
st.set_page_config(page_title="Vendas Diarias", layout="wide")  # ✅ Escolha um título só

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import re
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import plotly.express as px
import io
from st_aggrid import AgGrid, GridOptionsBuilder
from datetime import datetime, date
from datetime import datetime, date, timedelta
from calendar import monthrange


#st.set_page_config(page_title="Painel Agrupado", layout="wide")
#st.set_page_config(page_title="Vendas Diarias", layout="wide")
# 🔒 Bloqueia o acesso caso o usuário não esteja logado
if not st.session_state.get("acesso_liberado"):
    st.stop()
import streamlit as st

# =====================================
# CSS para esconder barra de botões do canto superior direito
# =====================================
st.markdown("""
    <style>
        [data-testid="stToolbar"] {
            visibility: hidden;
            height: 0%;
            position: fixed;
        }
    </style>
""", unsafe_allow_html=True)
# ================================
# 1. Conexão com Google Sheets
# ================================
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
credentials_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
gc = gspread.authorize(credentials)
planilha_empresa = gc.open("Vendas diarias")



df_empresa = pd.DataFrame(planilha_empresa.worksheet("Tabela Empresa").get_all_records())

# ================================
# 2. Configuração inicial do app
# ================================


# 🎨 Estilizar abas
st.markdown("""
    <style>
    .stApp { background-color: #f9f9f9; }
    div[data-baseweb="tab-list"] { margin-top: 20px; }
    button[data-baseweb="tab"] {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 10px 20px;
        margin-right: 10px;
        transition: all 0.3s ease;
        font-size: 16px;
        font-weight: 600;
    }
    button[data-baseweb="tab"]:hover { background-color: #dce0ea; color: black; }
    button[data-baseweb="tab"][aria-selected="true"] { background-color: #0366d6; color: white; }
    </style>
""", unsafe_allow_html=True)

# Cabeçalho bonito
st.markdown("""
    <div style='display: flex; align-items: center; gap: 10px; margin-bottom: 20px;'>
        <img src='https://img.icons8.com/color/48/graph.png' width='40'/>
        <h1 style='display: inline; margin: 0; font-size: 2.4rem;'>Relatórios</h1>
    </div>
""", unsafe_allow_html=True)

# ================================
# 3. Separação em ABAS
# ================================
aba1, aba3, aba4, aba5 = st.tabs([
    "📈 Gráficos",
    "📆 Relatórios Vendas",
    "📋 Relatório Diario Vendas/Metas",
    "📋 Relatórios Financeiros"
])
# ================================
# Aba 1: Graficos Anuais
# ================================
with aba1:
    planilha = gc.open("Vendas diarias")
    aba = planilha.worksheet("Fat Sistema Externo")
    dados = aba.get_all_records()
    df = pd.DataFrame(dados)

      
    # ✅ Limpa espaços invisíveis nos nomes das colunas
    df.columns = df.columns.str.strip()
    
    #st.write("🧪 Colunas carregadas:", df.columns.tolist())
    
    
   
    def limpar_valor(x):
        try:
            if isinstance(x, str):
                return float(x.replace("R$", "").replace(".", "").replace(",", ".").strip())
            return float(x)
        except:
            return None

    for col in ["Fat.Total", "Serv/Tx", "Fat.Real"]:
        if col in df.columns:
            df[col] = df[col].apply(limpar_valor)

    # Datas e campos derivados
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
    df["Ano"] = df["Data"].dt.year
    df["Mês"] = df["Data"].dt.month
    meses_portugues = {
        1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
        7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"
    }
    df["Nome Mês"] = df["Mês"].map(meses_portugues)
    
    # 0) Segurança: checa colunas mínimas
    cols_min = {"Data","Ano","Fat.Total","Loja"}
    if not cols_min.issubset(df.columns):
        st.error("❌ A planilha precisa ter as colunas: Data, Ano, Fat.Total, Loja.")
        st.stop()
    
    # ===== Filtros lado a lado (Ano | Filtrar por | Selecionar | Tipo) =====
    col_ano, col_tipo_filtro, col_alvo, col_tipo = st.columns([1.2, 1.0, 2.2, 1.2])
    
    with col_ano:
        anos_disponiveis = sorted(df["Ano"].dropna().unique())
        default_anos = anos_disponiveis[-2:] if len(anos_disponiveis) >= 2 else anos_disponiveis
        anos_comparacao = st.multiselect(
            "Ano",
            options=anos_disponiveis,
            default=default_anos
        )
    
    # 1) Filtra o dataframe pelos anos escolhidos
    df_anos = df[df["Ano"].isin(anos_comparacao)].dropna(subset=["Data","Fat.Total"]).copy()
    
    # 2) Adiciona "Tipo" via Tabela Empresa (se existir)
    try:
        df_emp = df_empresa[["Loja","Tipo"]].copy()
    except NameError:
        df_emp = pd.DataFrame(columns=["Loja","Tipo"])
    
    df_emp["Loja_norm"]  = df_emp["Loja"].astype(str).str.strip().str.lower()
    df_anos["Loja_norm"] = df_anos["Loja"].astype(str).str.strip().str.lower()
    df_anos = df_anos.merge(df_emp[["Loja_norm","Tipo"]], on="Loja_norm", how="left")
    df_anos.drop(columns=["Loja_norm"], inplace=True)
    df_anos["Tipo"] = df_anos["Tipo"].fillna("Sem tipo")
    
    # 3) “Filtrar por” (Loja ou Grupo) – só mostra o que existir
    opcoes_filtrar = [x for x in ["Loja","Grupo"] if x in df_anos.columns]
    with col_tipo_filtro:
        tipo_filtro = st.selectbox("Filtrar por", opcoes_filtrar, index=0)
    
    # 4) Multiselect dependente (pode escolher várias Lojas/Grupos)
    with col_alvo:
        opcoes_alvo = sorted(df_anos[tipo_filtro].astype(str).str.strip().dropna().unique())
        selecoes_alvo = st.multiselect(
            f"Selecionar {tipo_filtro}(s)",
            options=opcoes_alvo,
            default=[],  # vazio = todas
            help="Deixe em branco para considerar todas."
        )
    
    # 5) Filtro de Tipo (opcional)
    with col_tipo:
        tipos_opts = ["Todos"] + sorted(df_anos["Tipo"].dropna().astype(str).unique())
        tipo_selecionado = st.selectbox("Tipo", options=tipos_opts, index=0)
    
    # 6) Aplica filtros escolhidos
    if selecoes_alvo:
        df_anos = df_anos[df_anos[tipo_filtro].isin(selecoes_alvo)]
    if tipo_selecionado != "Todos":
        df_anos = df_anos[df_anos["Tipo"] == tipo_selecionado]
    
    # 7) Contagem de lojas únicas por ano (normalizada)
    _aux = df_anos.copy()
    _aux["Loja_norm"] = _aux["Loja"].astype(str).str.strip().str.lower()
    df_lojas = (_aux.drop_duplicates(subset=["Ano","Loja_norm"])
                  .groupby("Ano")["Loja_norm"].nunique()
                  .reset_index()
                  .rename(columns={"Loja_norm":"Qtd_Lojas"}))



   
    
    # Calcular a quantidade de lojas únicas por ano (com base em loja + ano únicos)
    df_lojas = df_anos.drop_duplicates(subset=["Ano", "Loja"])
    df_lojas = df_lojas.groupby("Ano")["Loja"].nunique().reset_index()
    df_lojas.columns = ["Ano", "Qtd_Lojas"]

     

    fat_mensal = df_anos.groupby(["Nome Mês", "Ano"])["Fat.Total"].sum().reset_index()

    meses = {
        "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
        "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12
    }
    fat_mensal["MesNum"] = fat_mensal["Nome Mês"].str[:3].str.lower().map(meses)
    fat_mensal["Ano"] = fat_mensal["Ano"].astype(str)
    fat_mensal["MesAno"] = fat_mensal["Nome Mês"].str[:3].str.capitalize() + "/" + fat_mensal["Ano"].str[-2:]
    fat_mensal = fat_mensal.sort_values(["MesNum", "Ano"])


    # Paleta de cores pastel fixa (azul, cinza e amarelo claro)
    cores_pastel = [
        "#A3C4F3",  # azul pastel
        "#BFD7FF",  # azul claro pastel
        "#E2E8F0",  # cinza azulado claro
        "#E5E7EB",  # cinza claro
        "#FFF3B0",  # amarelo pastel
        "#FDF6B2",  # amarelo claro
    ]
    
    # Mapeia cada ano para uma cor pastel automaticamente
    anos_presentes = sorted(fat_mensal["Ano"].astype(str).unique())
    color_map = {ano: cores_pastel[i % len(cores_pastel)] for i, ano in enumerate(anos_presentes)}


    # ===================================================
    # 🎨 Mapeamento fixo de cores por posição do ano
    # Último ano = cinza, penúltimo = amarelo, antepenúltimo = azul
    # Repete o ciclo para anos mais antigos
    # ===================================================
    
    # Ordena anos em ordem crescente
    anos_presentes = sorted(fat_mensal["Ano"].astype(str).unique())
    
    # Cores fixas (em ordem de prioridade: último → cinza, penúltimo → amarelo, antepenúltimo → azul)
    cores_ciclo = [
        "#08810E",  # cinza claro
        "#5CA0B7",  # amarelo pastel
        "#A3C4F3",  # azul pastel
    ]
    
    # Cria um mapeamento ano → cor seguindo o padrão
    color_map = {}
    for i, ano in enumerate(reversed(anos_presentes)):  
        cor = cores_ciclo[i % len(cores_ciclo)]  
        color_map[ano] = cor
    
    # Inverte novamente para que fique ano mais antigo → cor correspondente
    color_map = {ano: color_map[ano] for ano in anos_presentes}



    
    # ================================
    # 📊 Faturamento Mensal — Barras (ano mais recente) + Linhas (até 2 comparativos)
    # ================================
    import plotly.graph_objects as go
    
    # ordem fixa dos meses
    ordem_meses = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                   "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    
    # garante todos os meses na ordem e preenche faltantes com 0
    def series_por_ano(ano_str):
        base = pd.DataFrame({"Nome Mês": ordem_meses})
        s = fat_mensal[fat_mensal["Ano"] == str(ano_str)][["Nome Mês","Fat.Total"]]
        return base.merge(s, on="Nome Mês", how="left").fillna({"Fat.Total": 0})
    
    # formatação em milhões
    def fmt_mi(x):
        return f"{x/1_000_000:.1f} Mi"
    
    # anos disponíveis no filtro atual
    anos_sel = sorted(fat_mensal["Ano"].astype(int).unique())
    if not anos_sel:
        st.warning("Sem dados para os anos selecionados.")
        fig = go.Figure()
    else:
        ano_barras = max(anos_sel)  # ano atual (barras)
        # até 2 comparativos mais recentes para linhas
        comparativos = sorted([a for a in anos_sel if a != ano_barras], reverse=True)[:2]
    
        # dados do ano das barras
        df_bar = series_por_ano(ano_barras)
        cor_barra_padrao = color_map.get(str(ano_barras), "#A3C4F3")
    
        # último mês presente no ano atual e "até dd/mm"
        mes_idx = {m: i for i, m in enumerate(ordem_meses)}
        mes_atual_serie = fat_mensal.loc[fat_mensal["Ano"] == str(ano_barras), "Nome Mês"]
        ultimo_mes = max(mes_atual_serie, key=lambda m: mes_idx.get(m, -1)) if not mes_atual_serie.empty else None
        abbr = {"Janeiro":"JAN","Fevereiro":"FEV","Março":"MAR","Abril":"ABR","Maio":"MAI","Junho":"JUN",
                "Julho":"JUL","Agosto":"AGO","Setembro":"SET","Outubro":"OUT","Novembro":"NOV","Dezembro":"DEZ"}
    
        # melhor pegar a data real do DF original filtrado
        ultimo_registro = df_anos.loc[
            (df_anos["Ano"] == ano_barras) & (df_anos["Nome Mês"] == ultimo_mes), "Data"
        ].max()
        dia_lbl = ultimo_registro.strftime("%d/%m") if pd.notnull(ultimo_registro) else ""
    
        # rótulos das barras: mês atual com 2 linhas (valor + "até dd/mm")
        labels_barras = [
            fmt_mi(v) if m != ultimo_mes else f"{fmt_mi(v)}<br><b>até {dia_lbl}</b>"
            for m, v in zip(df_bar["Nome Mês"], df_bar["Fat.Total"])
        ]
    
        # (opcional) hatch no mês atual para indicar parcial
        pattern_shapes = ["/" if m == ultimo_mes else "" for m in df_bar["Nome Mês"]]
    
        # barra vermelha se menor que o comparativo mais recente (se existir)
        if comparativos:
            df_comp = series_por_ano(comparativos[0])
            cores_barras = [
                "#FF0000" if vb < vc else cor_barra_padrao
                for vb, vc in zip(df_bar["Fat.Total"].tolist(), df_comp["Fat.Total"].tolist())
            ]
        else:
            cores_barras = [cor_barra_padrao] * len(df_bar)
    
        fig = go.Figure()
    
        # barras do ano mais recente
        fig.add_trace(go.Bar(
            x=df_bar["Nome Mês"],
            y=df_bar["Fat.Total"],
            name=str(ano_barras),
            marker=dict(
                color=cores_barras,
                pattern=dict(shape=pattern_shapes, fgcolor="rgba(0,0,0,0.45)", solidity=0.25)  # remova se não quiser hatch
            ),
            opacity=0.95,
            text=labels_barras,
            textposition="outside",
            textfont=dict(color="black", size=12, family="Arial Black, Arial, sans-serif")
        ))
    
        # linhas: comparativo mais recente = vermelho, anterior = amarelo
        cores_linhas_fixas = ["#FF0000", "#FFD54F"]  # vermelho, amarelo pastel
        for idx, ano_l in enumerate(comparativos):
            df_lin = series_por_ano(ano_l)
            cor_linha = cores_linhas_fixas[idx % len(cores_linhas_fixas)]
    
            fig.add_trace(go.Scatter(
                x=df_lin["Nome Mês"],
                y=df_lin["Fat.Total"],
                name=str(ano_l),
                mode="lines+markers+text",
                line=dict(color=cor_linha, width=4),
                marker=dict(size=6, color=cor_linha),
                text=[fmt_mi(v) for v in df_lin["Fat.Total"]],
                textposition="top center",
                textfont=dict(size=12, color="black", family="Arial Black, Arial, sans-serif"),
                hoverinfo="skip"
            ))
    
            # rótulo do ano no início da linha (à esquerda de Janeiro)
            y0 = float(df_lin.loc[df_lin["Nome Mês"] == "Janeiro", "Fat.Total"].iloc[0])
            fig.add_annotation(
                x="Janeiro", y=y0,
                text=str(ano_l),
                showarrow=False,
                xanchor="right", yanchor="bottom",
                xshift=-60,  # ajuste fino
                font=dict(color=cor_linha, size=12, family="Arial", weight="bold")
            )
    
            fig.add_annotation(
                xref="paper", yref="paper", x=0, y=1.12,
                text=f"{ano_barras} parcial (acumulado até {dia_lbl}). Linhas = anos completos.",
                showarrow=False, font=dict(size=12, color="#555")
            )
            
        # layout final
        fig.update_layout(
            template="simple_white",
            xaxis_title=None,
            yaxis_title=None,
            xaxis=dict(tickangle=-45, domain=[0.02, 1]),  # espaço p/ rótulo do ano
            showlegend=False,
            margin=dict(t=10, b=10, l=0, r=0),
            paper_bgcolor="white",
            plot_bgcolor="white"
        )
    
    # ================================
    # 📊 Faturamento Anual — Horizontal
    # ================================
    df_total = fat_mensal.groupby("Ano")["Fat.Total"].sum().reset_index()
    df_total["Ano"] = df_total["Ano"].astype(int)
    df_lojas["Ano"] = df_lojas["Ano"].astype(int)
    df_total = df_total.merge(df_lojas, on="Ano", how="left")
    df_total["AnoTexto"] = df_total.apply(
        lambda row: f"{int(row['Ano'])}     R$ {row['Fat.Total']/1_000_000:,.1f} Mi".replace(",", "."), axis=1
    )
    
      
    # ordem correta
    anos_ordenados = sorted(df_total["Ano"].unique())
    anos_ordenados_str = [str(ano) for ano in anos_ordenados]
    df_total["Ano"] = pd.Categorical(df_total["Ano"].astype(str), categories=anos_ordenados_str, ordered=True)
    df_total = df_total.sort_values("Ano", ascending=True)
    
    fig_total = px.bar(
        df_total,
        x="Fat.Total",
        y="Ano",
        orientation="h",
        color="Ano",
        color_discrete_map=color_map
    )
    
    for _, row in df_total.iterrows():
        fig_total.add_annotation(
            x=0.1, y=row["Ano"], text=row["AnoTexto"],
            showarrow=False, xanchor="left", yanchor="middle",
            font=dict(color="black", size=16, family="Arial", weight="bold"),
            xref="x", yref="y"
        )
        fig_total.add_annotation(
            x=row["Fat.Total"], y=row["Ano"],
            showarrow=False, text=f"{int(row['Qtd_Lojas'])} Lojas",
            xanchor="left", yanchor="bottom", yshift=-8,
            font=dict(color="red", size=16, family="Arial", weight="bold"),
            xref="x", yref="y"
        )
    
    fig_total.update_layout(
        height=130, margin=dict(t=0, b=0, l=0, r=0),
        xaxis=dict(visible=False),
        yaxis=dict(categoryorder="array", categoryarray=anos_ordenados_str,
                   showticklabels=False, showgrid=False, zeroline=False),
        yaxis_title=None, showlegend=False,
        plot_bgcolor="rgba(0,0,0,0)"
    )
    
    # ----- título dinâmico p/ Faturamento Anual -----
    from datetime import date
    import numpy as np
    
    # último dia registrado do ano das barras (ano_barras)
    ultimo_registro = pd.to_datetime(
        df_anos.loc[df_anos["Ano"] == ano_barras, "Data"].max()
    )
    dia_lbl = ultimo_registro.strftime("%d/%m") if pd.notnull(ultimo_registro) else ""
    
    # mostra o sufixo apenas quando for comparativo e ano em curso
    comparando = len(comparativos) > 0
    ano_em_curso = (int(ano_barras) == date.today().year)
    mostrar_parcial = comparando and ano_em_curso and bool(dia_lbl)
    
    titulo_anual = "Faturamento Anual"
    if mostrar_parcial:
        titulo_anual += f" ({ano_barras} até {dia_lbl})"
    
    # =================== RENDER: ANUAL NO TOPO (FULL WIDTH) ===================
    fig_total.update_layout(
        margin=dict(l=0, r=0, t=10, b=10),
        xaxis=dict(domain=[0.0, 1.0])  # ocupa toda a largura
    )
    st.subheader(titulo_anual)
    st.plotly_chart(fig_total, use_container_width=True, theme=None)
    
    # ============== CSS para tabelas compactas (fonte e padding menores) ==============
    st.markdown("""
    <style>
    /* Fonte e altura menores apenas na tabela de Participação */
    div[data-testid="stDataFrame"][aria-label*="Participação Faturamento"] table {
        font-size: 9px !important;
    }
    div[data-testid="stDataFrame"][aria-label*="Participação Faturamento"] thead tr th {
        padding: 0px 3px !important;
    }
    div[data-testid="stDataFrame"][aria-label*="Participação Faturamento"] div[role="row"] {
        min-height: 16px !important;
    }
    div[data-testid="stDataFrame"][aria-label*="Participação Faturamento"] div[role="gridcell"] {
        padding: 0px 3px !important;
        line-height: 1 !important;
    }
    </style>
    """, unsafe_allow_html=True)
        
    # =================== LINHA DE BAIXO: MENSAL (ESQ) + TABELAS (DIR) ===================
    col_chart, col_tables = st.columns([0.80, 0.20])  # ajuste a proporção se quiser
    
    with col_chart:
        # “cola” o mensal mais à esquerda
        fig.update_layout(
            margin=dict(l=0, r=0, t=10, b=10),
            xaxis=dict(domain=[0.06, 0.98])
        )
        st.subheader("Faturamento Mensal")
        st.plotly_chart(fig, use_container_width=True, theme=None)
    
   
    with col_tables:
        # ---------- Helpers ----------
        meses_idx = {m: i+1 for i, m in enumerate(ordem_meses)}
        m_lim = meses_idx.get(ultimo_mes, 12) if ultimo_mes else 12
        
        ano_atual = int(ano_barras)          # <- último ano filtrado
        ano_prev  = ano_atual - 1
        
        # Recorte YTD (até o último mês do ano atual)
        ytd = df_anos[df_anos["Mês"] <= m_lim].copy()
        
        # Usa "Operação" se existir; senão "Grupo"; senão "Loja"
        dim = "Operação" if "Operação" in ytd.columns else ("Grupo" if "Grupo" in ytd.columns else "Loja")
        
        # Função para encurtar nomes (melhora o espaçamento)
        def encurta(txt, limite=10):
            s = str(txt).strip()
            return (s[:limite] + "…") if len(s) > limite else s
        
        # =========================
        # 1) OPERAÇÃO: Qtd de lojas por ano (YTD)  (mantém 2 colunas: prev e atual)
        # =========================
        tmp = ytd.copy()
        tmp["Loja_norm"] = tmp["Loja"].astype(str).str.strip().str.lower()
        tab_op = (tmp.drop_duplicates(subset=["Ano", dim, "Loja_norm"])
                     .groupby(["Ano", dim])["Loja_norm"].nunique()
                     .unstack("Ano")
                     .reindex(columns=[ano_prev, ano_atual], fill_value=0)
                     .reset_index())
        
        tab_op.columns = ["OPERAÇÃO", str(ano_prev), str(ano_atual)]
        tab_op_exibe = tab_op.copy()
        tab_op_exibe["OPERAÇÃO"] = tab_op_exibe["OPERAÇÃO"].map(lambda x: encurta(x, 10))
        
        st.markdown("##### Operação")
        st.dataframe(tab_op_exibe, hide_index=True, use_container_width=True, height=120)
        
        # =========================
        # 2) Índice de Crescimento (YTD) — SEMPRE baseado no último ano filtrado
        #    Crescimento % = (YTD[ano_atual] / YTD[ano_prev] - 1)*100
        # =========================
        fat_ytd = (ytd.groupby(["Ano", dim])["Fat.Total"].sum()
                     .unstack("Ano")
                     .reindex(columns=[ano_prev, ano_atual], fill_value=0.0))
        
        cresc = fat_ytd.copy()
        cresc["Crescimento %"] = (
            (cresc[ano_atual] / cresc[ano_prev]).replace([np.inf, -np.inf], np.nan) - 1.0
        ) * 100
        
        cresc = (cresc[["Crescimento %"]]
                 .reset_index()
                 .sort_values("Crescimento %", ascending=False)
                 .fillna(0.0))
        
        cresc.rename(columns={dim: "OPERAÇÃO"}, inplace=True)
        cresc["OPERAÇÃO"] = cresc["OPERAÇÃO"].map(lambda x: encurta(x, 10))
        cresc["Crescimento %"] = cresc["Crescimento %"].map(
            lambda v: f"{v:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        )
        
        st.markdown("Crescimento")
        st.dataframe(cresc, hide_index=True, use_container_width=True, height=120)
        
        # =========================
        # 3) Participação Faturamento (YTD) — SEMPRE do último ano filtrado
        # =========================
        part = (ytd[ytd["Ano"] == ano_atual]
                  .groupby(dim)["Fat.Total"].sum()
                  .reset_index()
                  .rename(columns={"Fat.Total": "Faturamento"}))
        
        total_atual = part["Faturamento"].sum()
        part["Participação"] = (part["Faturamento"] / total_atual * 100).fillna(0.0)
        part = part.sort_values("Participação", ascending=False)
        
        part_exibe = part[[dim, "Participação"]].rename(columns={dim: "OPERAÇÃO"}).copy()
        part_exibe["OPERAÇÃO"] = part_exibe["OPERAÇÃO"].map(lambda x: encurta(x, 10))
        part_exibe["Participação"] = part_exibe["Participação"].map(
            lambda v: f"{v:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        )
        
        st.markdown("##### %Faturamento")
        st.dataframe(part_exibe, hide_index=True, use_container_width=True, height=120)
        
        # ---- CSS para compactar essas tabelas (fonte/padding/linhas) ----
        st.markdown("""
        <style>
        /* Fonte menor e menos padding nas tabelas renderizadas nesta coluna */
        div[data-testid="stDataFrame"] table { font-size: 9px; }
        div[data-testid="stDataFrame"] thead tr th { padding: 0 3px !important; }
        div[data-testid="stDataFrame"] div[role="gridcell"] { padding: 0 3px !important; line-height: 1 !important; }
        /* Títulos menores e com menos margem */
        h5 { margin: 6px 0 4px 0 !important; }
        </style>
        """, unsafe_allow_html=True)


# ================================
# Aba 3: Relatórios Vendas
# ================================
with aba3:
    import pandas as pd
    import numpy as np
    import streamlit as st
    from datetime import datetime

    st.markdown("""
    <style>
    /* Remove aparência de "caixa" do chip selecionado */
    .stMultiSelect [data-baseweb="tag"] {
        background-color: transparent !important;
        color: black !important;
        font-weight: 500 !important;
        border: none !important;
        box-shadow: none !important;
        padding: 2px 6px !important;
        margin: 2px 4px !important;
    }
    
    /* Remove o fundo da área de seleção */
    .stMultiSelect > div {
        background-color: transparent !important;
        border: none !important;
    }
    </style>
    """, unsafe_allow_html=True)


    # Carrega dados
    df_empresa = pd.DataFrame(planilha_empresa.worksheet("Tabela Empresa").get_all_records())
    df_vendas = pd.DataFrame(planilha_empresa.worksheet("Fat Sistema Externo").get_all_records())

    # Normalização
    df_empresa["Loja"] = df_empresa["Loja"].str.strip().str.upper()
    df_empresa["Grupo"] = df_empresa["Grupo"].str.strip()
    df_vendas.columns = df_vendas.columns.str.strip()
    df_vendas["Data"] = pd.to_datetime(df_vendas["Data"], dayfirst=True, errors="coerce")
    df_vendas["Loja"] = df_vendas["Loja"].astype(str).str.strip().str.upper()
    df_vendas["Grupo"] = df_vendas["Grupo"].astype(str).str.strip()

    # Merge com Tipo
    df_vendas = df_vendas.merge(
        df_empresa[["Loja", "Tipo"]],
        on="Loja",
        how="left"
    )    
    df_vendas["Fat.Total"] = (
        df_vendas["Fat.Total"]
        .astype(str)
        .str.replace("R$", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df_vendas["Fat.Total"] = pd.to_numeric(df_vendas["Fat.Total"], errors="coerce")

    # ==== Filtros principais ====
    # ==== Filtros principais ====
    data_min = df_vendas["Data"].min()
    data_max = df_vendas["Data"].max()
    
    col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
    
    with col1:
        modo_exibicao = st.selectbox("🔀 Ver por:", ["Loja", "Grupo"], key="modo_exibicao_relatorio")
    
    # Filtro de Loja ou Grupo (apenas define, será aplicado depois)
    todos = "Todas"
    opcoes_filtro = sorted(df_vendas[modo_exibicao].dropna().unique())
    opcoes_filtro.insert(0, todos)
    
    with col2:
        selecao = st.selectbox(
            f"🎯 Selecione {modo_exibicao.lower()} (opcional):",
            options=opcoes_filtro,
            index=0,
            key="filtro_unico_loja_grupo"
        )
    
    with col3:
        modo_periodo = st.selectbox("🕒 Período:", ["Diário", "Mensal", "Anual"], key="modo_periodo_relatorio")

    with col4:
        tipos_disponiveis = sorted(df_vendas["Tipo"].dropna().unique())
        tipos_disponiveis.insert(0, "Todos")
        tipo_selecionado = st.selectbox("🏪 Tipo:", options=tipos_disponiveis, index=0)

    
    
    # ==== Filtro por período ====
    # ==== Filtro por período ====
    if modo_periodo == "Diário":
        # 📅 Intervalo com validação
            datas_selecionadas = st.date_input(
                "📅 Intervalo de datas:",
                value=(data_max, data_max),
                min_value=data_min,
                max_value=data_max,
                key="data_vendas_relatorio"
            )
        
            if isinstance(datas_selecionadas, (tuple, list)) and len(datas_selecionadas) == 2:
                data_inicio, data_fim = datas_selecionadas
            else:
                st.warning("⚠️ Por favor, selecione um intervalo com **duas datas** (início e fim).")
                st.stop()
        
            df_filtrado = df_vendas[
                (df_vendas["Data"] >= pd.to_datetime(data_inicio)) &
                (df_vendas["Data"] <= pd.to_datetime(data_fim))
            ]
            # ==== Aplica filtro de Tipo ====
            if tipo_selecionado != "Todos":
                df_filtrado = df_filtrado[df_filtrado["Tipo"] == tipo_selecionado]

            df_filtrado["Período"] = df_filtrado["Data"].dt.strftime("%d/%m/%Y")

    
    elif modo_periodo == "Mensal":
        df_vendas["Mes/Ano"] = df_vendas["Data"].dt.strftime("%m/%Y")
        meses_disponiveis = sorted(
            df_vendas["Mes/Ano"].unique(),
            key=lambda x: datetime.strptime("01/" + x, "%d/%m/%Y")
        )
        meses_selecionados = st.multiselect(
            "🗓️ Selecione os meses:",
            options=meses_disponiveis,
            default=[datetime.today().strftime("%m/%Y")]
        )
        df_filtrado = df_vendas[df_vendas["Mes/Ano"].isin(meses_selecionados)]
        df_filtrado["Período"] = df_filtrado["Data"].dt.strftime("%m/%Y")

        if tipo_selecionado != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Tipo"] == tipo_selecionado]

    
    elif modo_periodo == "Anual":
        df_vendas["Ano"] = df_vendas["Data"].dt.strftime("%Y")
        anos_disponiveis = sorted(df_vendas["Ano"].unique())
        anos_selecionados = st.multiselect(
            "📅 Selecione os anos:",
            options=anos_disponiveis,
            default=[datetime.today().strftime("%Y")]
        )
        df_filtrado = df_vendas[df_vendas["Ano"].isin(anos_selecionados)]
        df_filtrado["Período"] = df_filtrado["Data"].dt.strftime("%Y")

        if tipo_selecionado != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Tipo"] == tipo_selecionado]

    
    # ==== Aplica filtro de Loja ou Grupo ====
    if selecao != todos:
        df_filtrado = df_filtrado[df_filtrado[modo_exibicao] == selecao]

    
        

    # Agrupamento
    chaves = ["Tipo", "Loja", "Grupo"] if modo_exibicao == "Loja" else ["Tipo", "Grupo"]
    df_agrupado = df_filtrado.groupby(chaves + ["Período"], as_index=False)["Fat.Total"].sum()

    # Pivot
    df_pivot = df_agrupado.pivot_table(index=chaves, columns="Período", values="Fat.Total", fill_value=0).reset_index()

    # Ordenação das datas
    def ordenar_datas(col):
        try:
            return datetime.strptime(col, "%d/%m/%Y")
        except:
            try:
                return datetime.strptime("01/" + col, "%d/%m/%Y")
            except:
                return datetime.strptime("01/01/" + col, "%d/%m/%Y")

    colunas_periodo = sorted(
        [c for c in df_pivot.columns if c not in ["Tipo", "Loja", "Grupo"]],
        key=ordenar_datas
    )

    # Garante que colunas existam
    if "Loja" not in df_pivot.columns:
        df_pivot["Loja"] = ""
    if "Grupo" not in df_pivot.columns:
        df_pivot["Grupo"] = ""

    # Define ordem final: Grupo, Loja, depois períodos
    colunas_finais = ["Tipo", "Grupo", "Loja"] + colunas_periodo
    df_final = df_pivot[colunas_finais].copy()

    # Total acumulado (última coluna)
    ultima_coluna_valor = colunas_periodo[-1]
    df_final["__ordem"] = df_final[ultima_coluna_valor]

    # Ordena do maior para o menor
    df_final = df_final.sort_values(by="__ordem", ascending=False).drop(columns="__ordem").reset_index(drop=True)

    # === Linha de Lojas Ativas (quantas lojas venderam algo em cada período) ===
    # === Linha de Lojas Ativas (quantas lojas venderam por período) ===
    df_lojas_por_periodo = df_filtrado.groupby("Período")["Loja"].nunique()

    # 🔢 Soma total de todas as colunas de valor
    colunas_valor = colunas_periodo  # já está definido corretamente
    soma_total_geral = df_final[colunas_valor].sum(numeric_only=True).sum()
    
    # ➕ Cria uma coluna auxiliar com o total por linha
    df_final["__soma_linha"] = df_final[colunas_valor].sum(axis=1, numeric_only=True)
    
    # 🎯 Calcula o percentual total com base na soma de todas as colunas
    df_final["% Total"] = df_final["__soma_linha"] / soma_total_geral
    
    # 🧹 Remove a auxiliar
    df_final = df_final.drop(columns=["__soma_linha"])


    
    # Monta linha com estrutura do df_final, já convertendo os valores para str (evita float!)
    linha_lojas = {col: "" for col in df_final.columns}
    linha_lojas["Grupo"] = "Lojas Ativas"
    linha_lojas["Loja"] = ""
    linha_lojas["Tipo"] = ""
        
    for periodo in df_lojas_por_periodo.index:
        if periodo in linha_lojas:
            linha_lojas[periodo] = str(int(df_lojas_por_periodo[periodo]))  # 👈 força int e depois str
    
    # === Linha TOTAL ===
    linha_total = df_final.drop(columns=["Grupo", "Loja", "Tipo"]).sum(numeric_only=True)
    linha_total["Grupo"] = "TOTAL"
    linha_total["Loja"] = ""
    linha_total["Tipo"] = ""
    
    # Junta tudo: Lojas Ativas → TOTAL → Dados
    df_final = pd.concat([
        pd.DataFrame([linha_lojas]),
        pd.DataFrame([linha_total]),
        df_final
    ], ignore_index=True)

    # Formatação
    def formatar(valor):
        try:
            return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return valor

    # === Formatação final ===
    df_formatado = df_final.copy()
    
    for col in colunas_periodo:
        if col in df_formatado.columns:
            # Só aplica formatação em valores monetários (não em 'Lojas Ativas')
            df_formatado[col] = df_formatado.apply(
                lambda row: formatar(row[col]) if row["Grupo"] not in ["Lojas Ativas"] else row[col],
                axis=1
            )
    df_formatado = df_formatado[["Tipo", "Grupo", "Loja"] + colunas_periodo + ["% Total"]]
    #df_formatado = df_formatado[["Tipo", "Grupo", "Loja"] + colunas_periodo]
    # Formata a nova coluna de percentual
    df_percentual = pd.to_numeric(df_final["% Total"], errors="coerce")
    df_formatado["% Total"] = df_percentual.apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "")

    # Estilo para destacar TOTAL
    def aplicar_estilo(df):
        def estilo_linha(row):
            if row["Grupo"] == "TOTAL":
                return ["background-color: #f0f0f0; font-weight: bold"] * len(row)
            elif row["Grupo"] == "Lojas Ativas":
                return ["background-color: #eeeeee; font-style: italic"] * len(row)  # ⛔ aqui está o destaque
            else:
                return ["" for _ in row]
        return df.style.apply(estilo_linha, axis=1)

    # Se for exibição por Grupo, remove a coluna "Loja" antes de aplicar estilo
    df_visivel = df_formatado.drop(columns=["Loja"]) if modo_exibicao == "Grupo" else df_formatado
    
    # Aplica estilo na nova tabela visual
    tabela_final = aplicar_estilo(df_visivel)
    
    # Exibe
    st.dataframe(
        tabela_final,
        use_container_width=True,
        height=700
    )
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # ➤ Usa a mesma lógica da visualização: remove "Loja" se for modo Grupo
    df_exportar = df_final.drop(columns=["Loja"]) if modo_exibicao == "Grupo" else df_final.copy()
    
    # ➤ Exporta para Excel com BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exportar.to_excel(writer, index=False, sheet_name="Relatório")
    output.seek(0)
    
    # ➤ Reabre com openpyxl para aplicar formatação
    wb = load_workbook(output)
    ws = wb["Relatório"]
    
    # === Estilos ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")  # Azul escuro
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # ➤ Aplica estilo no cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # ➤ Detecta nome da última coluna (geralmente "% Total")
    ultima_coluna_nome = df_exportar.columns[-1]
    
    # ➤ Aplica formatação e estilo nas células
    # ➤ Aplica formatação e estilo nas células
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        grupo_valor = row[1].value  # Coluna "Grupo" (segunda coluna)
        estilo_fundo = None
    
        if isinstance(grupo_valor, str):
            if grupo_valor.strip().upper() == "TOTAL":
                estilo_fundo = PatternFill("solid", fgColor="F4B084")  # Laranja escuro
            elif grupo_valor.strip().upper() == "LOJAS ATIVAS":
                estilo_fundo = PatternFill("solid", fgColor="D9D9D9")  # Cinza claro
    
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment
            if estilo_fundo:
                cell.fill = estilo_fundo
    
            col_name = ws.cell(row=1, column=cell.column).value  # Nome da coluna
    
            if isinstance(cell.value, (int, float)):
                if col_name == "% Total":
                    cell.number_format = "0.000%"
                else:
                    cell.number_format = '"R$" #,##0.00'

    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    # ➤ Ajusta a largura das colunas automaticamente
    from openpyxl.utils import get_column_letter

    # ➤ Ajusta a largura das colunas com base no valor formatado
    for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    cell_str = str(cell.value)
                    max_length = max(max_length, len(cell_str))
            except:
                pass
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 2




    # ➤ Alinha à esquerda as colunas "Tipo", "Grupo" e "Loja"
    colunas_df = list(df_exportar.columns)
    colunas_esquerda = ["Tipo", "Grupo", "Loja"]
    for col_nome in colunas_esquerda:
        if col_nome in colunas_df:
            col_idx = colunas_df.index(col_nome) + 1
            for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for c in cell:
                    c.alignment = Alignment(horizontal="left")
    
    # ➤ Define largura fixa da coluna "Loja" para 22
    if "Loja" in colunas_df:
        col_idx_loja = colunas_df.index("Loja")
        col_letra_loja = chr(ord("A") + col_idx_loja)
        ws.column_dimensions[col_letra_loja].width = 22


    
    # ➤ Salva para download
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    
    # ➤ Botão de download
    st.download_button(
        label="📥 Baixar Excel",
        data=output_final,
        file_name="Relatorio_Vendas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================================
# Aba 4: Relatório Vendas/Metas
# ================================
with aba4:
    # Carrega dados
    df_empresa = pd.DataFrame(planilha_empresa.worksheet("Tabela Empresa").get_all_records())
    df_vendas = pd.DataFrame(planilha_empresa.worksheet("Fat Sistema Externo").get_all_records())
    df_empresa["Loja"] = df_empresa["Loja"].str.strip().str.upper()
    df_empresa["Grupo"] = df_empresa["Grupo"].str.strip()
    df_vendas.columns = df_vendas.columns.str.strip()
    df_vendas["Data"] = pd.to_datetime(df_vendas["Data"], dayfirst=True, errors="coerce")
    df_vendas["Loja"] = df_vendas["Loja"].astype(str).str.strip().str.upper()
    df_vendas["Grupo"] = df_vendas["Grupo"].astype(str).str.strip()
    df_vendas["Fat.Total"] = (
        df_vendas["Fat.Total"]
        .astype(str)
        .str.replace("R$", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df_vendas["Fat.Total"] = pd.to_numeric(df_vendas["Fat.Total"], errors="coerce")
    
    
    
    # Filtros
    data_min = df_vendas["Data"].min()
    data_max = df_vendas["Data"].max()
    col1, col2, col3 = st.columns([2, 2, 2])
    ultimo_dia_disponivel = df_vendas["Data"].max()
    data_min = df_vendas["Data"].min()
    data_max = df_vendas["Data"].max()
    
    with col1:
        datas_selecionadas = st.date_input(
            "📅 Intervalo de datas:",
            value=(data_max, data_max),
            min_value=data_min,
            max_value=data_max
        )
        
        # Validação para garantir que foram selecionadas 2 datas
        if isinstance(datas_selecionadas, (tuple, list)) and len(datas_selecionadas) == 2:
            data_inicio, data_fim = datas_selecionadas

            # ⚠️ Bloqueia seleção entre meses diferentes
            if data_inicio.month != data_fim.month or data_inicio.year != data_fim.year:
                st.warning("⚠️ Selecione datas **dentro do mesmo mês** e **ano**. Não é permitido misturar meses no intervalo.")
                st.stop()
        else:
            st.warning("⚠️ Por favor, selecione um intervalo com **duas datas** (início e fim).")
            st.stop()



       
    with col2:
        modo_exibicao = st.selectbox("🧭 Ver por:", ["Loja", "Grupo"])
    
    # -----------------------------------
    # 🎛️ Seletor dinâmico de colunas extras (sem dependência de "Meta" ou "Sem Meta")
    # -----------------------------------
    st.markdown("### 🎛️ Personalize sua visualização")
    
    colunas_opcionais = {
        "🎯 Meta da Loja": "Meta",
        "📊 % Atingido": "%Atingido",
        "🏬 % Loja X Operação": "%LojaXGrupo",
        "🧮 % Operação no Total": "%Grupo"
    }
    
    opcoes_selecionadas = st.multiselect(
        "➕ Escolha os indicadores que deseja **exibir**:",
        options=list(colunas_opcionais.keys()),
        default=["🎯 Meta da Loja", "📊 % Atingido", "🧮 % Operação no Total"]
    )
    
    # Mapeia as escolhas visuais para os nomes reais das colunas
    colunas_escolhidas = [colunas_opcionais[op] for op in opcoes_selecionadas]

    
    data_inicio_dt = pd.to_datetime(data_inicio)
    data_fim_dt = pd.to_datetime(data_fim)
    primeiro_dia_mes = data_fim_dt.replace(day=1)
    datas_periodo = pd.date_range(start=data_inicio_dt, end=data_fim_dt)
    
    # Base combinada com 0s
    df_lojas_grupos = df_empresa[["Loja", "Grupo"]].drop_duplicates()
    df_base_completa = pd.MultiIndex.from_product(
        [df_lojas_grupos["Loja"], datas_periodo], names=["Loja", "Data"]
    ).to_frame(index=False)
    df_base_completa = df_base_completa.merge(df_lojas_grupos, on="Loja", how="left")
    df_filtro_dias = df_vendas[(df_vendas["Data"] >= data_inicio_dt) & (df_vendas["Data"] <= data_fim_dt)]
    df_agrupado_dias = df_filtro_dias.groupby(["Data", "Loja", "Grupo"], as_index=False)["Fat.Total"].sum()
    df_completo = df_base_completa.merge(df_agrupado_dias, on=["Data", "Loja", "Grupo"], how="left")
    df_completo["Fat.Total"] = df_completo["Fat.Total"].fillna(0)
    
    # Pivot com datas
    df_pivot = df_completo.pivot_table(
        index=["Grupo", "Loja"], columns="Data", values="Fat.Total", aggfunc="sum", fill_value=0
    ).reset_index()
    df_pivot.columns = [
        col if isinstance(col, str) else f"Fat Total {col.strftime('%d/%m/%Y')}"
        for col in df_pivot.columns
    ]
    
    # Acumulado do mês
    df_mes = df_vendas[(df_vendas["Data"] >= primeiro_dia_mes) & (df_vendas["Data"] <= data_fim_dt)]
    df_acumulado = df_mes.groupby(["Loja", "Grupo"], as_index=False)["Fat.Total"].sum()
    df_acumulado = df_lojas_grupos.merge(df_acumulado, on=["Loja", "Grupo"], how="left")
    df_acumulado["Fat.Total"] = df_acumulado["Fat.Total"].fillna(0)
    col_acumulado = f"Acumulado Mês (01/{data_fim_dt.strftime('%m')} até {data_fim_dt.strftime('%d/%m')})"
    df_acumulado = df_acumulado.rename(columns={"Fat.Total": col_acumulado})
    df_base = df_pivot.merge(df_acumulado, on=["Grupo", "Loja"], how="left")
    df_base = df_base[df_base[col_acumulado] != 0]
    
    # Adiciona coluna de Meta
    df_metas = pd.DataFrame(planilha_empresa.worksheet("Metas").get_all_records())
    df_metas["Loja"] = df_metas["Loja Vendas"].astype(str).str.strip().str.upper()
    mapa_meses = {
        "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04", "MAI": "05", "JUN": "06",
        "JUL": "07", "AGO": "08", "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
    }
    df_metas["Mês"] = df_metas["Mês"].astype(str).str.strip().str.upper().map(mapa_meses)
    df_metas["Ano"] = df_metas["Ano"].astype(str).str.strip()
    df_metas["Meta"] = (
        df_metas["Meta"]
        .astype(str)
        .str.replace("R$", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df_metas["Meta"] = pd.to_numeric(df_metas["Meta"], errors="coerce").fillna(0)
    mes_filtro = data_fim_dt.strftime("%m")
    ano_filtro = data_fim_dt.strftime("%Y")
    df_metas_filtrado = df_metas[(df_metas["Mês"] == mes_filtro) & (df_metas["Ano"] == ano_filtro)].copy()
    df_base["Loja"] = df_base["Loja"].astype(str).str.strip().str.upper()
    # Adiciona coluna Meta
    df_base = df_base.merge(df_metas_filtrado[["Loja", "Meta"]], on="Loja", how="left")
    
    # Adiciona coluna Tipo (vindo de Tabela Empresa)
    # Merge da coluna Tipo
    df_base = df_base.merge(
        df_empresa[["Loja", "Tipo","PDV"]].drop_duplicates(), 
        on="Loja", 
        how="left", 
        validate="many_to_one"
    )
    
    
    df_base["Meta"] = df_base["Meta"].fillna(0)
    
    
    # %Atingido
    df_base["%Atingido"] = df_base[col_acumulado] / df_base["Meta"]
    df_base["%Atingido"] = df_base["%Atingido"].replace([np.inf, -np.inf], np.nan).fillna(0).round(4)
    
    # Reordena colunas
    colunas_base = ["Grupo", "PDV", "Loja", "Tipo"]
    from datetime import datetime
    
    col_diarias = [
        col for col in df_base.columns if col.startswith("Fat Total")
    ]
    
    # Extrai a data do nome da coluna e ordena corretamente
    col_diarias.sort(key=lambda x: datetime.strptime(x.replace("Fat Total ", ""), "%d/%m/%Y"))
    #colunas_finais = colunas_base + col_diarias + [col_acumulado, "Meta", "%Atingido", "%LojaXGrupo", "%Grupo"]
    colunas_finais = colunas_base + col_diarias + [col_acumulado] + colunas_escolhidas
    for col in ["%LojaXGrupo", "%Grupo"]:
        if col not in df_base.columns:
            df_base[col] = np.nan
    # Garante que Tipo está presente antes de selecionar colunas finais
    if "Tipo" not in df_base.columns:
        df_base = df_base.merge(df_empresa[["Loja", "Tipo"]].drop_duplicates(), on="Loja", how="left")
    
    # 🔧 Define colunas visíveis antes de qualquer concatenação
    df_base = df_base[colunas_finais]
    colunas_visiveis = colunas_finais.copy()
    
    # 🔢 Linha total
    linha_total = df_base.drop(columns=["Grupo", "PDV", "Loja", "Tipo"]).sum(numeric_only=True)
    linha_total["Grupo"] = "TOTAL"
    linha_total["PDV"] = df_base.drop_duplicates(subset="Loja")["PDV"].sum()  # não faz sentido somar PDV
    linha_total["Loja"] = f"Lojas: {df_base['Loja'].nunique():02d}"
    linha_total["Tipo"] = ""
    
    # 🧱 Agrupa por grupo
    # 🧱 Agrupa por grupo e define o Tipo do grupo
    ordem_tipos = ["Airports", "Kopp", "On-Premise"]
    ordem_tipo_dict = {tipo: i for i, tipo in enumerate(ordem_tipos)}
    
    grupos_info = []
    for grupo, df_grp in df_base.groupby("Grupo"):
        df_grp = df_grp.copy()
        
        # Detecta o tipo mais comum do grupo (ou NA se indefinido)
        tipo_dominante = df_grp["Tipo"].dropna().mode().iloc[0] if not df_grp["Tipo"].dropna().empty else "—"
        tipo_ordenado = ordem_tipo_dict.get(tipo_dominante, 999)
        
        total_grupo = df_grp[col_acumulado].sum()
        grupos_info.append((tipo_ordenado, grupo, total_grupo, df_grp, tipo_dominante))
    
    # 📊 Ordena primeiro por Tipo, depois por acumulado (decrescente)
    grupos_info.sort(key=lambda x: (x[0], -x[2]))
    
    
    # 🔁 Monta blocos
    blocos = []
    for _, grupo, _, df_grp, tipo_dominante in grupos_info:
        df_grp_ord = df_grp.sort_values(by=col_acumulado, ascending=False)
    
        # Subtotal
        tipo_valor = tipo_dominante
    
        subtotal = df_grp_ord.drop(columns=["Grupo", "Loja", "Tipo"]).sum(numeric_only=True)
        subtotal["Grupo"] = f"{'SUBTOTAL ' if modo_exibicao == 'Loja' else ''}{grupo}"
        subtotal["PDV"] = df_grp_ord.drop_duplicates(subset="Loja")["PDV"].sum()
        subtotal["Loja"] = f"Lojas: {df_grp_ord['Loja'].nunique():02d}"
        subtotal["Tipo"] = tipo_valor
    
        # ✅ Garante todas as colunas
        for col in colunas_visiveis:
            if col not in subtotal:
                subtotal[col] = np.nan
        subtotal = subtotal[colunas_visiveis]
    
        # 🟦 Lojas
        if modo_exibicao == "Loja":
            blocos.append(df_grp_ord[colunas_visiveis])
    
        # 🟨 Subtotal
        blocos.append(pd.DataFrame([subtotal], columns=colunas_visiveis))
    
    # 🔚 Junta tudo
    linha_total = pd.DataFrame([linha_total], columns=colunas_visiveis)
    df_final = pd.concat([linha_total] + blocos, ignore_index=True)
    
    #st.write("🔍 Diagnóstico: Linhas de loja sem Tipo", df_final[(df_final["Tipo"].isna()) & (~df_final["Loja"].str.startswith("Lojas:"))])
    # Percentuais
    filtro_lojas = (
        (df_final["Loja"] != "") &
        (~df_final["Grupo"].astype(str).str.startswith("SUBTOTAL")) &
        (df_final["Grupo"] != "TOTAL")
    )
    df_lojas_reais = df_final[filtro_lojas].copy()
    soma_por_grupo = df_lojas_reais.groupby("Grupo")[col_acumulado].transform("sum")
    soma_total_geral = df_lojas_reais[col_acumulado].sum()
    if modo_exibicao != "Grupo":
        df_final.loc[filtro_lojas, "%LojaXGrupo"] = (
            df_lojas_reais[col_acumulado].values / soma_por_grupo.values
        ).round(4)
    if modo_exibicao == "Grupo":
        filtro_grupos = df_final["Loja"].astype(str).str.startswith("Lojas:")
    else:
        filtro_grupos = df_final["Grupo"].astype(str).str.startswith("SUBTOTAL")
    if "%Grupo" in colunas_escolhidas:
        df_final.loc[filtro_grupos, "%Grupo"] = (
            df_final.loc[filtro_grupos, col_acumulado] / soma_total_geral
        ).round(4)
    
    # %Atingido final
    if modo_exibicao == "Loja":
        mascara_subtotal = df_final["Grupo"].astype(str).str.startswith("SUBTOTAL")
        mascara_total = df_final["Grupo"] == "TOTAL"
        df_final.loc[mascara_subtotal | mascara_total, "%Atingido"] = (
            df_final.loc[mascara_subtotal | mascara_total, col_acumulado] /
            df_final.loc[mascara_subtotal | mascara_total, "Meta"]
        ).replace([np.inf, -np.inf], np.nan).fillna(0).round(4)
    else:
        mascara_total = df_final["Grupo"] == "TOTAL"
        mascara_subtotal = df_final["Loja"].astype(str).str.startswith("Lojas:")
        df_final.loc[mascara_total | mascara_subtotal, "%Atingido"] = (
            df_final.loc[mascara_total | mascara_subtotal, col_acumulado] /
            df_final.loc[mascara_total | mascara_subtotal, "Meta"]
        ).replace([np.inf, -np.inf], np.nan).fillna(0).round(4)
        df_final.loc[~(mascara_total | mascara_subtotal), "%Atingido"] = ""
    
    # Oculta coluna %LojaXGrupo se for modo Grupo
    # Define colunas com base no filtro "Meta" ou "Sem Meta"
    colunas_visiveis = ["Grupo", "PDV", "Loja", "Tipo"] + col_diarias + [col_acumulado] + colunas_escolhidas
    if "Tipo" in colunas_visiveis:
        colunas_visiveis.remove("Tipo")
    df_final = df_final[colunas_visiveis]
    
    # Formata valores
    colunas_percentuais = ["%LojaXGrupo", "%Grupo", "%Atingido"]
    def formatar(valor, col):
        try:
            if pd.isna(valor) or valor == "":
                return ""
            return f"{valor:.2%}".replace(".", ",") if col in colunas_percentuais else f"R$ {valor:,.2f}".replace(".", "X").replace(",", ".").replace("X", ",")
        except:
            return ""
    # Soma total da meta do mês filtrado
    total_meta_mes = df_metas_filtrado["Meta"].sum()
    
    df_formatado = df_final.copy()
    for col in colunas_visiveis:
        if col in colunas_percentuais:
            df_formatado[col] = df_formatado[col].apply(lambda x: formatar(x, col))
        elif col not in ["Grupo", "Loja", "Tipo", "PDV"]:
            df_formatado[col] = df_formatado[col].apply(lambda x: formatar(x, col))
        else:
            df_formatado[col] = df_formatado[col].fillna("")  # 👈 tipo e loja não numéricos
    
    # ================================
    # ➕ Linhas de resumo por Tipo
    # ================================
    df_base_tipo = df_base.copy()
    
    # Ignora lojas sem tipo
    df_base_tipo = df_base_tipo[~df_base_tipo["Tipo"].isna()]
    
    linhas_resumo_tipo = []
    tipos_ordenados = df_base_tipo.groupby("Tipo")[col_acumulado].sum().sort_values(ascending=False).index.tolist()
    
    for tipo in tipos_ordenados:
        df_tipo_filtro = df_base_tipo[df_base_tipo["Tipo"] == tipo]
        if df_tipo_filtro.empty:
            continue
    
        linha = {}
        linha["Grupo"] = tipo
        linha["Loja"] = f"Lojas: {df_tipo_filtro['Loja'].nunique():02d}"
        linha["Tipo"] = tipo if pd.notna(tipo) and tipo != "" else "—"  # 👈 aqui
        linha["PDV"] = df_tipo_filtro.drop_duplicates(subset="Loja")["PDV"].sum()
    
        # Somatórios
        for col in col_diarias:
            linha[col] = df_tipo_filtro[col].sum()
    
        linha[col_acumulado] = df_tipo_filtro[col_acumulado].sum()
    
        # Meta total por tipo
        linha["Meta"] = df_tipo_filtro["Meta"].sum()

        # % Atingido
        linha["%Atingido"] = (
            linha[col_acumulado] / linha["Meta"]
            if linha["Meta"] != 0 else 0
        )

        

        # %Grupo no total
        soma_total_geral = df_base_tipo[col_acumulado].sum()
        if "%Grupo" in colunas_escolhidas:
            linha["%Grupo"] = ( 
                linha[col_acumulado] / soma_total_geral
                if soma_total_geral != 0 else 0
            )
    
        linhas_resumo_tipo.append(linha)
    
    
    df_resumo_tipo = pd.DataFrame(linhas_resumo_tipo)
    
    # Formata
    df_resumo_tipo_formatado = df_resumo_tipo.copy()
    #for col in df_resumo_tipo.columns:
    #    if col not in ["Grupo", "Loja"]:
    #        df_resumo_tipo_formatado[col] = df_resumo_tipo[col].apply(lambda x: formatar(x, col))

    for col in df_resumo_tipo.columns:
        if col not in ["Grupo", "Loja"]:
            if col == "PDV":
                # Mantém como inteiro (sem R$)
                df_resumo_tipo_formatado[col] = df_resumo_tipo[col].astype(int).astype(str)
            else:
                df_resumo_tipo_formatado[col] = df_resumo_tipo[col].apply(lambda x: formatar(x, col))
        
    
    
    
    
    
    
    
    
    
    
    # Calcula o percentual desejável até o dia selecionado
    dia_hoje = data_fim_dt.day
    dias_mes = monthrange(data_fim_dt.year, data_fim_dt.month)[1]
    perc_desejavel = dia_hoje / dias_mes
    
    # Cálculo proporcional da meta até o dia
    meta_total_mes = df_metas_filtrado["Meta"].sum()
    meta_desejada_ate_hoje = (meta_total_mes / dias_mes) * dia_hoje
    
    # Faturamento desejável (com ordem correta das colunas)
    # Faturamento desejável (com ordem correta das colunas)
    linha_desejavel_dict = {}
    for col in colunas_visiveis:
        if col == "Grupo":
            linha_desejavel_dict[col] = ""
        elif col == "Loja":
            linha_desejavel_dict[col] = f"FATURAMENTO IDEAL ATÉ {data_fim_dt.strftime('%d/%m')}"
        elif col == "%Atingido":
            linha_desejavel_dict[col] = formatar(perc_desejavel, "%Atingido")
        elif col == col_acumulado:
            linha_desejavel_dict[col] = formatar(meta_desejada_ate_hoje, "Meta")
        else:
            linha_desejavel_dict[col] = ""
    
    linha_desejavel = pd.DataFrame([linha_desejavel_dict])
    
    # 🔽 Remove "Tipo" da visualização final
    for df_temp in [df_resumo_tipo_formatado, df_formatado, linha_desejavel]:
        if "Tipo" in df_temp.columns:
            df_temp.drop(columns=["Tipo"], inplace=True)
    
    # 🔁 Junta tudo para exibir
    df_linhas_visiveis = pd.concat([df_resumo_tipo_formatado, df_formatado], ignore_index=True)
    df_exibir = pd.concat([linha_desejavel, df_linhas_visiveis], ignore_index=True)
    
    # 🎨 Define função de estilo
    def aplicar_estilo_final(df, estilos_linha):
        def apply_row_style(row):
            base_style = estilos_linha[row.name].copy()
            if "%Atingido" in df.columns and row.name > 0:
                try:
                    valor = row["%Atingido"]
                    if isinstance(valor, str) and "%" in valor:
                        valor_float = float(valor.replace("%", "").replace(",", ".")) / 100
                    else:
                        valor_float = float(valor)
                    if not pd.isna(valor_float):
                        idx = df.columns.get_loc("%Atingido")
                        if valor_float >= perc_desejavel:
                            base_style[idx] += "; color: green; font-weight: bold"
                        else:
                            base_style[idx] += "; color: red; font-weight: bold"
                except:
                    pass
            return base_style
        return df.style.apply(apply_row_style, axis=1)
    
    # 🎨 Estilo visual por linha
    cores_alternadas = ["#eef4fa", "#f5fbf3"]
    estilos_linha = []
    cor_idx = -1
    grupo_atual = None
    
    tem_grupo_resumo = (
        'df_resumo_tipo_formatado' in locals()
        and not df_resumo_tipo_formatado.empty
        and "Grupo" in df_resumo_tipo_formatado.columns
    )
    
    for _, row in df_linhas_visiveis.iterrows():
        grupo = row["Grupo"]
        loja = row["Loja"]
    
        if isinstance(grupo, str) and tem_grupo_resumo and grupo in df_resumo_tipo_formatado["Grupo"].values:
            estilos_linha.append(["background-color: #fffbea; font-weight: bold"] * len(row))
        elif grupo == "TOTAL":
            estilos_linha.append(["background-color: #f2f2f2; font-weight: bold"] * len(row))
        elif isinstance(grupo, str) and grupo.startswith("SUBTOTAL"):
            estilos_linha.append(["background-color: #fff8dc; font-weight: bold"] * len(row))
        elif loja == "":
            estilos_linha.append(["background-color: #fdfdfd"] * len(row))
        else:
            if grupo != grupo_atual:
                cor_idx = (cor_idx + 1) % len(cores_alternadas)
                grupo_atual = grupo
            cor = cores_alternadas[cor_idx]
            estilos_linha.append([f"background-color: {cor}"] * len(row))
    
    # ➕ Linha desejável no topo
    estilos_final = [["background-color: #dddddd; font-weight: bold"] * len(df_linhas_visiveis.columns)]
    estilos_final += estilos_linha
    # 🔁 Renomeia coluna 'Grupo' apenas visualmente para 'Operação'
    df_exibir.columns = [col.replace("Grupo", "Operação") for col in df_exibir.columns]
    # 📊 Exibe resultado final
    st.dataframe(
        aplicar_estilo_final(df_exibir, estilos_final),
        use_container_width=True,
        height=750
    )
    
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side
    
    # Gera o Excel já na memória
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"
    
    # Define bordas
    border_padrao = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    border_grossa = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    
    # Cabeçalho com azul escuro
    # === Linha "Faturamento Desejável" na LINHA 1 ===
    # --- Linha FATURAMENTO DESEJÁVEL na linha 1 ---
    for col_idx, (col_nome, valor) in enumerate(linha_desejavel.iloc[0].items(), start=1):
        if col_nome.strip().upper() == "PDV":
            valor = ""  # ⛔ não exibe valor na linha Faturamento Ideal
        cell = ws.cell(row=1, column=col_idx, value=valor)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")  # cinza claro
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left" if col_idx == 2 else "right")
    
        # 🔁 Aplica borda apenas nas laterais (remove linhas internas)
        if col_idx == 1:
            cell.border = Border(left=Side(style="thin"))
        elif col_idx == len(linha_desejavel.columns):
            cell.border = Border(right=Side(style="thin"))
        else:
            cell.border = Border()  # sem borda

    
    # === Cabeçalho na LINHA 2 ===
    for col_idx, col in enumerate(df_exibir.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col)
        cell.fill = PatternFill("solid", fgColor="1F4E78")  # azul escuro
        cell.font = Font(bold=True, color="FFFFFF")         # texto branco
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_grossa
    
    # Ajuste de altura das linhas
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30

    
    # Preenche os dados na planilha
    # Dentro do loop de preenchimento de dados
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Estilos de borda
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    border_padrao = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_grossa = Border(left=thick, right=thick, top=thick, bottom=thick)

    if "PDV" in df_exibir.columns:
        df_exibir["PDV"] = pd.to_numeric(df_exibir["PDV"], errors="coerce").fillna(0).astype(int)
    
    
    # Dados
    # Detecta total de linhas que serão preenchidas
    # Preenche os dados na planilha
    for row_idx, (i, row) in enumerate(df_exibir.iterrows(), start=3):
        estilo_linha = estilos_final[row_idx - 3]
    
        grupo = row.get("Operação", "")
        grupo_str = str(grupo).strip().upper()  # garante consistência
        is_subtotal = grupo_str.startswith("SUBTOTAL")
        is_total = grupo_str == "TOTAL"
        usar_borda_grossa = is_subtotal or is_total
    
        for col_idx, (col, valor) in enumerate(row.items(), start=1):

            # 🎯 Trate "PDV" como número inteiro SEMPRE, com prioridade total
            if col.strip().upper() == "PDV":
                try:
                    valor_int = int(str(valor).strip().replace(".0", ""))
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor_int)
                    cell.number_format = '0'  # inteiro puro
                except:
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        
            elif isinstance(valor, str) and "%" in valor:
                try:
                    valor_float = float(valor.replace("%", "").replace(",", ".")) / 100
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor_float)
                    cell.number_format = '0.00%'
                except:
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        
            elif isinstance(valor, str) and "R$" in valor:
                try:
                    valor_float = float(valor.replace("R$", "").replace(".", "").replace(",", "."))
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor_float)
                    cell.number_format = 'R$ #,##0.00'
                except:
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)


    
           # 🎨 Estilo de fundo com prioridade ao TOTAL e SUBTOTAL
           # 🎨 Estilo de fundo com prioridade ao TOTAL e SUBTOTAL
            estilo = estilo_linha[col_idx - 1]
            cor_personalizada = None
            
            if is_total:
                cor_personalizada = "F4B183"  # Laranja escuro p/ TOTAL
            elif is_subtotal:
                cor_personalizada = "FCE4D6"  # Laranja claro p/ SUBTOTAL
            elif "background-color" in estilo:
                cor_personalizada = estilo.split("background-color: ")[1].split(";")[0].replace("#", "")
            
            if cor_personalizada:
                cell.fill = PatternFill("solid", fgColor=cor_personalizada)

                
            # 🅱️ Negrito
            if "font-weight: bold" in estilo:
                cell.font = Font(bold=True)
    
            # 📏 Alinhamento
            # Detecta a posição da coluna 'Loja'
            pos_coluna_loja = list(df_exibir.columns).index("Loja")
            
            # Define alinhamento: esquerda até 'Loja', centralizado depois
            if col_idx - 1 <= pos_coluna_loja:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
    
            # 🧱 Bordas
            if row_idx == 3:  # Linha Faturamento Ideal Desejável
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))
            else:
                cell.border = border_grossa if usar_borda_grossa else border_padrao
    
            # ✅ Cor verde/vermelha no %Atingido
            if col == "%Atingido":
                try:
                    if isinstance(valor, str) and "%" in valor:
                        valor_float = float(valor.replace("%", "").replace(",", ".")) / 100
                    elif isinstance(valor, (int, float)):
                        valor_float = float(valor)
                    else:
                        valor_float = None
    
                    if valor_float is not None:
                        if valor_float >= perc_desejavel:
                            cell.font = Font(color="006400", bold=True)  # Verde escuro
                        else:
                            cell.font = Font(color="B22222", bold=True)  # Vermelho escuro
                except:
                    pass



    
        
    
        
    
    
    # ⬇️ Ajusta automaticamente a largura das colunas
    # Ajuste refinado de largura das colunas
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.number_format == 'R$ #,##0.00' and isinstance(cell.value, (float, int)):
                    cell_str = f'R$ {cell.value:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")
                    length = len(cell_str)
                elif cell.number_format == '0.00%' and isinstance(cell.value, (float, int)):
                    cell_str = f'{cell.value:.2%}'.replace(".", ",")
                    length = len(cell_str)
                else:
                    cell_str = str(cell.value) if cell.value is not None else ""
                    length = len(cell_str)
                max_length = max(max_length, length)
            except:
                pass
    
        adjusted_width = max_length + 2  # margem extra
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width
    # 🔥 Exclui a linha duplicada "FATURAMENTO IDEAL ATÉ..." da planilha final
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):  # pula cabeçalhos
        cell_val = row[2].value  # coluna B → índice 1
        if isinstance(cell_val, str) and cell_val.startswith("FATURAMENTO IDEAL ATÉ"):
            ws.delete_rows(row[0].row, 1)
            break  # remove apenas a primeira ocorrência

    
    # Salva em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Botão direto de download
    st.download_button(
        label="📥 Baixar Excel",
        data=buffer,
        file_name="vendas_formatado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ======================
# 📝 Relatórios Financeiros
# ======================

with aba5:
    try:
        st.markdown("""
        <div style="background-color:#fff3cd; border-left: 6px solid #ffecb5; padding: 1rem; border-radius: 6px; font-size: 16px;">
        🚧 <strong>Este relatório está em desenvolvimento.</strong> Resultados e funcionalidades podem mudar a qualquer momento.
        </div>
        """, unsafe_allow_html=True)
        
        # Carrega a planilha (caso ainda não tenha feito antes)
        planilha = gc.open("Vendas diarias")

        # Aba com dados analíticos
        aba_relatorio = planilha.worksheet("Faturamento Meio Pagamento")
        df_relatorio = pd.DataFrame(aba_relatorio.get_all_records())
        df_relatorio.columns = df_relatorio.columns.str.strip()

        # Aba com o tipo de pagamento
        aba_meio_pagamento = planilha.worksheet("Tabela Meio Pagamento")
        df_meio_pagamento = pd.DataFrame(aba_meio_pagamento.get_all_records())
        df_meio_pagamento.columns = df_meio_pagamento.columns.str.strip()

        # Normaliza colunas usadas no merge
        df_relatorio["Meio de Pagamento"] = df_relatorio["Meio de Pagamento"].astype(str).str.strip().str.upper()
        df_meio_pagamento["Meio de Pagamento"] = df_meio_pagamento["Meio de Pagamento"].astype(str).str.strip().str.upper()
        df_meio_pagamento["Tipo de Pagamento"] = df_meio_pagamento["Tipo de Pagamento"].astype(str).str.strip().str.upper()

        # Merge para adicionar "Tipo de Pagamento"
        df_relatorio = df_relatorio.merge(
            df_meio_pagamento[["Meio de Pagamento", "Tipo de Pagamento"]],
            on="Meio de Pagamento",
            how="left"
        )

        # Corrige valores e datas
        df_relatorio["Valor (R$)"] = (
            df_relatorio["Valor (R$)"].astype(str)
            .str.replace("R$", "", regex=False)
            .str.replace("(", "-")
            .str.replace(")", "")
            .str.replace(" ", "")
            .str.replace(".", "")
            .str.replace(",", ".")
            .astype(float)
        )
        df_relatorio["Data"] = pd.to_datetime(df_relatorio["Data"], dayfirst=True, errors="coerce")

        # Datas mínimas e máximas
        data_min = df_relatorio["Data"].min().date()
        data_max = df_relatorio["Data"].max().date()

        # ===== FILTROS GERAIS =====
        # 📅 Seleção de data
        datas_selecionadas = st.date_input(
            "📅 Intervalo de datas:",
            value=(data_max, data_max),
            min_value=data_min,
            max_value=data_max
        )
        if isinstance(datas_selecionadas, (tuple, list)) and len(datas_selecionadas) == 2:
            data_inicio, data_fim = datas_selecionadas
        else:
            st.warning("⚠️ Por favor, selecione um intervalo com **duas datas** (início e fim).")
            st.stop()
        
        # 💳 Filtro abaixo da data (não mais ao lado)
        tipos_disponiveis = df_relatorio["Tipo de Pagamento"].dropna().unique().tolist()
        tipos_disponiveis.sort()
        filtro_tipo_pagamento = st.multiselect(
            "💳 Tipo de Pagamento:",
            options=tipos_disponiveis,
            default=tipos_disponiveis
        )
        
        # 🔎 Aplica filtro global
        df_filtrado = df_relatorio[
            (df_relatorio["Data"].dt.date >= data_inicio) &
            (df_relatorio["Data"].dt.date <= data_fim) &
            (df_relatorio["Tipo de Pagamento"].isin(filtro_tipo_pagamento))
        ]
        
        if df_filtrado.empty:
            st.info("🔍 Não há dados para o período e filtros selecionados.")
            st.stop()


        # ====== TABS ======
        aba_vendas, aba_taxas, aba_financeiro, aba_previsao_fc, aba_conciliacao = st.tabs([
            "💰 Vendas meio pagamento",
            "🔗 Vendas + Prazo e Taxas",
            "📄 Financeiro (Recebimentos)",
            "💰 Previsão FC",
            "🔄 Conciliação Adquirente"
        ])

        # === ABA VENDAS ===
        
        with aba_vendas:
                    
            agrupamento = st.radio(
                "Escolha como deseja agrupar os dados:",
                options=["Grupo", "Loja", "Meio de Pagamento", "Tipo de Pagamento"],
                horizontal=True
            )
        
            if agrupamento == "Grupo":
                index_cols = ["Grupo", "Meio de Pagamento"]
            elif agrupamento == "Loja":
                index_cols = ["Grupo", "Loja", "Meio de Pagamento"]  # Grupo sempre antes de Loja
            elif agrupamento == "Meio de Pagamento":
                index_cols = ["Meio de Pagamento"]
            elif agrupamento == "Tipo de Pagamento":
                index_cols = ["Tipo de Pagamento", "Meio de Pagamento"]
        
            # Cria a tabela dinâmica
            df_pivot = pd.pivot_table(
                df_filtrado,
                index=index_cols,
                columns=df_filtrado["Data"].dt.strftime("%d/%m/%Y"),
                values="Valor (R$)",
                aggfunc="sum",
                fill_value=0
            ).reset_index()
        
            # Renomeia colunas de data
            novo_nome_datas = {col: f"Vendas - {col}" for col in df_pivot.columns if "/" in str(col)}
            df_pivot.rename(columns=novo_nome_datas, inplace=True)
        
            # Calcula total por linha
            df_pivot["Total Vendas"] = df_pivot[[c for c in df_pivot.columns if "Vendas -" in c]].sum(axis=1)
        
            # Adiciona linha de TOTAL GERAL
            linha_total = {col: df_pivot[col].sum() if np.issubdtype(df_pivot[col].dtype, np.number) else "TOTAL GERAL" for col in df_pivot.columns}
            df_pivot_total = pd.concat([pd.DataFrame([linha_total]), df_pivot], ignore_index=True)
        
            # Formata números
            for col in df_pivot_total.select_dtypes(include=[np.number]).columns:
                df_pivot_total[col] = df_pivot_total[col].map(
                    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )
        
            st.dataframe(df_pivot_total, use_container_width=True)

        import xlsxwriter
        from io import BytesIO
        
        # === Prepara dados para exportar
        df_export = df_pivot_total.copy()
        
        # === Remove o símbolo R$ e converte para número para aplicar formato corretamente no Excel
        colunas_valores = [col for col in df_export.columns if "Vendas" in col or "Total Vendas" in col]
        for col in colunas_valores:
            df_export[col] = (
                df_export[col].astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df_export[col] = pd.to_numeric(df_export[col], errors="coerce").fillna(0.0)
        
        # === Gera arquivo Excel com formatação
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Vendas")
        
        # === Formatos
        formato_header = workbook.add_format({'bold': True, 'bg_color': '#CCCCCC'})
        formato_reais = workbook.add_format({'num_format': 'R$ #,##0.00'})
        
        # === Escreve cabeçalhos
        for col_idx, col_name in enumerate(df_export.columns):
            worksheet.write(0, col_idx, col_name, formato_header)
        
        # === Escreve dados com formatação de reais
        for row_idx, row in enumerate(df_export.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row):
                col_name = df_export.columns[col_idx]
                if col_name in colunas_valores:
                    worksheet.write_number(row_idx, col_idx, value, formato_reais)
                else:
                    worksheet.write(row_idx, col_idx, value)
        
        # === Ajusta larguras das colunas
        for i, col in enumerate(df_export.columns):
            max_width = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(i, i, max_width)
        
        workbook.close()
        output.seek(0)
        
        # === Botão de download
        st.download_button(
            label="⬇️ Baixar Excel",
            data=output,
            file_name="Vendas_Meio_Pagamento.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # === ABA PRAZO E TAXAS ===
        with aba_taxas:
            df_completo = df_filtrado.merge(
                df_meio_pagamento[["Meio de Pagamento", "Prazo", "Antecipa S/N", "Taxa Bandeira", "Taxa Antecipação"]],
                on="Meio de Pagamento",
                how="left"
            )

            df_pivot = pd.pivot_table(
                df_completo,
                index=["Meio de Pagamento", "Prazo", "Antecipa S/N", "Taxa Bandeira", "Taxa Antecipação"],
                columns=df_completo["Data"].dt.strftime("%d/%m/%Y"),
                values="Valor (R$)",
                aggfunc="sum",
                fill_value=0
            ).reset_index()

            colunas_datas = [col for col in df_pivot.columns if "/" in col]
            novo_nome_datas = {col: f"Vendas - {col}" for col in colunas_datas}
            df_pivot.rename(columns=novo_nome_datas, inplace=True)
            df_pivot.rename(columns={"Vendas - Antecipa S/N": "Antecipa S/N"}, inplace=True)

            colunas_vendas = [col for col in df_pivot.columns if "Vendas" in col]
            cols_fixas = ["Meio de Pagamento", "Prazo", "Antecipa S/N", "Taxa Bandeira", "Taxa Antecipação"]
            novas_cols = []

            for col_vendas in colunas_vendas:
                data_col = col_vendas.split(" - ")[1]
                col_taxa_bandeira = f"Vlr Taxa Bandeira - {data_col}"
                taxa_bandeira = (
                    pd.to_numeric(df_pivot["Taxa Bandeira"].astype(str)
                                  .str.replace("%","")
                                  .str.replace(",","."),
                                  errors="coerce").fillna(0) / 100
                )
                df_pivot[col_taxa_bandeira] = df_pivot[col_vendas] * taxa_bandeira

                col_taxa_antecipacao = f"Vlr Taxa Antecipação - {data_col}"
                taxa_antecipacao = (
                    pd.to_numeric(df_pivot["Taxa Antecipação"].astype(str)
                                  .str.replace("%","")
                                  .str.replace(",","."),
                                  errors="coerce").fillna(0) / 100
                )
                df_pivot[col_taxa_antecipacao] = df_pivot[col_vendas] * taxa_antecipacao

                novas_cols.extend([col_vendas, col_taxa_bandeira, col_taxa_antecipacao])

            df_pivot = df_pivot[cols_fixas + novas_cols]

            df_pivot["Total Vendas"] = df_pivot[colunas_vendas].sum(axis=1)
            df_pivot["Total Tx Bandeira"] = df_pivot[[col for col in df_pivot.columns if "Vlr Taxa Bandeira" in col]].sum(axis=1)
            df_pivot["Total Tx Antecipação"] = df_pivot[[col for col in df_pivot.columns if "Vlr Taxa Antecipação" in col]].sum(axis=1)
            df_pivot["Total a Receber"] = df_pivot["Total Vendas"] - df_pivot["Total Tx Bandeira"] - df_pivot["Total Tx Antecipação"]

            linha_total_dict = {col: "" for col in df_pivot.columns}
            linha_total_dict["Meio de Pagamento"] = "TOTAL GERAL"
            for col in df_pivot.columns:
                if "Vendas" in col or "Vlr Taxa Bandeira" in col or "Vlr Taxa Antecipação" in col \
                    or "Total Tx" in col or col in ["Total Vendas", "Total a Receber"]:
                    linha_total_dict[col] = df_pivot[col].sum()

            linha_total = pd.DataFrame([linha_total_dict])
            df_pivot_total = pd.concat([linha_total, df_pivot], ignore_index=True)

            df_pivot_exibe = df_pivot_total.copy()
            for col in [c for c in df_pivot_exibe.columns if "Vendas" in c or "Vlr Taxa Bandeira" in c 
                        or "Vlr Taxa Antecipação" in c or "Total Tx" in c or c in ["Total Vendas", "Total a Receber"]]:
                df_pivot_exibe[col] = df_pivot_exibe[col].map(
                    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )

            st.dataframe(df_pivot_exibe, use_container_width=True)


            import xlsxwriter
            from io import BytesIO
            
            # Cópia do DataFrame
            df_export = df_pivot_total.copy()
            
            # Conversões de dados
            df_export["Prazo"] = pd.to_numeric(df_export["Prazo"], errors="coerce").fillna(0).astype(int)
            df_export["Taxa Bandeira"] = (
                df_export["Taxa Bandeira"].astype(str)
                .str.replace("%", "").str.replace(",", ".")
            )
            df_export["Taxa Antecipação"] = (
                df_export["Taxa Antecipação"].astype(str)
                .str.replace("%", "").str.replace(",", ".")
            )
            
            df_export["Taxa Bandeira"] = pd.to_numeric(df_export["Taxa Bandeira"], errors="coerce").fillna(0.0) / 100
            df_export["Taxa Antecipação"] = pd.to_numeric(df_export["Taxa Antecipação"], errors="coerce").fillna(0.0) / 100
            
            # Colunas de valores reais
            # Colunas que devem receber formato de reais
            colunas_valores = [col for col in df_export.columns if any(p in col for p in [
                "Vendas", "Vlr Taxa Bandeira", "Vlr Taxa Antecipação", "Total"
            ])]
            
            # Geração do Excel
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet("Prazo e Taxas")
            
            # Formatos
            formato_header = workbook.add_format({'bold': True, 'bg_color': '#CCCCCC'})
            formato_reais = workbook.add_format({'num_format': 'R$ #,##0.00'})
            formato_percentual = workbook.add_format({'num_format': '0.00%'})
            formato_inteiro = workbook.add_format({'num_format': '0'})
            formato_texto = workbook.add_format()
            
            # Cabeçalhos
            for col_idx, col_name in enumerate(df_export.columns):
                worksheet.write(0, col_idx, col_name, formato_header)
            
            # Conteúdo
            for row_idx, row in enumerate(df_export.itertuples(index=False), start=1):
                for col_idx, value in enumerate(row):
                    col_name = df_export.columns[col_idx]
            
                    if col_name == "Prazo":
                        worksheet.write_number(row_idx, col_idx, value, formato_inteiro)
                    elif col_name in ["Taxa Bandeira", "Taxa Antecipação"]:
                        worksheet.write_number(row_idx, col_idx, value, formato_percentual)
                    elif col_name in colunas_valores:
                        worksheet.write_number(row_idx, col_idx, value, formato_reais)
                    else:
                        worksheet.write(row_idx, col_idx, str(value), formato_texto)
            
            # Ajusta largura
            for i, col in enumerate(df_export.columns):
                max_width = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, max_width)
            
            workbook.close()
            output.seek(0)
            
            # Botão de download
            st.download_button(
                label="⬇️ Baixar Excel",
                data=output,
                file_name="Prazo_Taxas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


        # === ABA FINANCEIRO ===
        with aba_financeiro:
            df_completo = df_filtrado.merge(
                df_meio_pagamento[["Meio de Pagamento", "Prazo", "Antecipa S/N"]],
                on="Meio de Pagamento",
                how="left"
            )
            df_completo["Prazo"] = pd.to_numeric(df_completo["Prazo"], errors="coerce").fillna(0).astype(int)
            df_completo["Antecipa S/N"] = df_completo["Antecipa S/N"].astype(str).str.upper().str.strip()

            from pandas.tseries.offsets import BDay
            df_completo["Data Recebimento"] = df_completo.apply(
                lambda row: row["Data"] + BDay(1) if row["Antecipa S/N"] == "SIM" else row["Data"] + BDay(row["Prazo"]),
                axis=1
            )

            df_financeiro = df_completo.groupby(df_completo["Data Recebimento"].dt.date)["Valor (R$)"].sum().reset_index()
            df_financeiro = df_financeiro.rename(columns={"Data Recebimento": "Data"})

            total_geral = df_financeiro["Valor (R$)"].sum()
            linha_total = pd.DataFrame([["TOTAL GERAL", total_geral]], columns=df_financeiro.columns)
            df_financeiro_total = pd.concat([linha_total, df_financeiro], ignore_index=True)

            df_financeiro_total["Valor (R$)"] = df_financeiro_total["Valor (R$)"].map(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )

            st.dataframe(df_financeiro_total, use_container_width=True)


    
            import xlsxwriter
            from io import BytesIO
            
            # === Corrige valores para float e converte a coluna Data para datetime
            df_export = df_financeiro_total.copy()
            
            # Trata "Valor (R$)"
            df_export["Valor (R$)"] = (
                df_export["Valor (R$)"]
                .astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df_export["Valor (R$)"] = pd.to_numeric(df_export["Valor (R$)"], errors="coerce").fillna(0.0)
            
            # Converte coluna "Data" para datetime (exceto TOTAL GERAL)
            df_export["Data"] = pd.to_datetime(df_export["Data"], errors="coerce")
            
            # Gera arquivo Excel com formatação
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet("Financeiro")
            
            # Formatos
            formato_header = workbook.add_format({'bold': True, 'bg_color': '#CCCCCC'})
            formato_reais = workbook.add_format({'num_format': 'R$ #,##0.00'})
            formato_data = workbook.add_format({'num_format': 'dd/mm/yyyy'})
            
            # Escreve cabeçalhos
            for col_idx, col_name in enumerate(df_export.columns):
                worksheet.write(0, col_idx, col_name, formato_header)
            
            # Escreve dados
            for row_idx, row in enumerate(df_export.itertuples(index=False), start=1):
                for col_idx, value in enumerate(row):
                    col_name = df_export.columns[col_idx]
                    if col_name == "Valor (R$)":
                        worksheet.write_number(row_idx, col_idx, value, formato_reais)
                    elif col_name == "Data":
                        if pd.notnull(value):
                            worksheet.write_datetime(row_idx, col_idx, value, formato_data)
                        else:
                            worksheet.write(row_idx, col_idx, "TOTAL GERAL")
                    else:
                        worksheet.write(row_idx, col_idx, value)
            
            # Ajusta largura das colunas
            for i, col in enumerate(df_export.columns):
                max_width = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, max_width)
            
            workbook.close()
            output.seek(0)
            
            # Botão de download
            st.download_button(
                label="⬇️ Baixar Excel",
                data=output,
                file_name="Financeiro Recebimento.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ========================
        # ========================
        # 📊 Aba Previsão FC
        # ========================
        with aba_previsao_fc:
            
            # Carrega planilha e abas
            planilha = gc.open("Vendas diarias")
            aba_fat = planilha.worksheet("Faturamento Meio Pagamento")
            aba_empresa = planilha.worksheet("Tabela Empresa")
        
            # --- Dados principais ---
            df_fat = pd.DataFrame(aba_fat.get_all_records())
            df_empresa = pd.DataFrame(aba_empresa.get_all_records())
        
            # --- Normalizações ---
            df_fat.columns = df_fat.columns.str.strip()
            df_empresa.columns = df_empresa.columns.str.strip()
        
            df_fat["Loja"] = df_fat["Loja"].astype(str).str.strip().str.upper()
            df_empresa["Loja"] = df_empresa["Loja"].astype(str).str.strip().str.upper()
            df_empresa["Grupo"] = df_empresa["Grupo"].astype(str).str.strip().str.upper()
        
            # Converte data
            df_fat["Data"] = pd.to_datetime(df_fat["Data"], dayfirst=True, errors="coerce")
            df_fat = df_fat.dropna(subset=["Data"])
        
            # Últimos 30 dias com base na aba correta
            data_final = df_fat["Data"].max()
            data_inicial = data_final - pd.Timedelta(days=30)
            df_30dias = df_fat[(df_fat["Data"] >= data_inicial) & (df_fat["Data"] <= data_final)].copy()
        
            # Traduz dia da semana
            dias_semana = {
                "Monday": "Segunda-feira",
                "Tuesday": "Terça-feira",
                "Wednesday": "Quarta-feira",
                "Thursday": "Quinta-feira",
                "Friday": "Sexta-feira",
                "Saturday": "Sábado",
                "Sunday": "Domingo"
            }
            df_30dias["Dia da Semana"] = df_30dias["Data"].dt.day_name().map(dias_semana)
        
            # Limpa e converte valores
            df_30dias["Valor (R$)"] = (
                df_30dias["Valor (R$)"]
                .astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df_30dias = df_30dias[df_30dias["Valor (R$)"].str.strip() != ""]
            df_30dias["Valor (R$)"] = pd.to_numeric(df_30dias["Valor (R$)"], errors="coerce")
            df_30dias = df_30dias.dropna(subset=["Valor (R$)"])
        
            # Seleciona colunas
            df_fc = df_30dias[[
                "Loja", "Data", "Dia da Semana", "Valor (R$)", 
                "Código Everest", "Código Grupo Everest"
            ]].copy()
        
            # Junta com Tipo e Grupo
            df_fc = df_fc.merge(df_empresa[["Loja", "Grupo", "Tipo"]], on="Loja", how="left")
        
            # Define ID FC
            def definir_id_fc(row):
                if row["Tipo"] == "Airports":
                    return row["Código Grupo Everest"]
                elif row["Tipo"] in ["Koop - Airports", "On-Premise"]:
                    return row["Código Everest"]
                else:
                    return None
        
            df_fc["ID FC"] = df_fc.apply(definir_id_fc, axis=1)
        
            # Agrupa e calcula média
            df_resultado = (
                df_fc.groupby(["Grupo", "Loja", "ID FC", "Dia da Semana"])["Valor (R$)"]
                .mean()
                .reset_index()
                .rename(columns={"Valor (R$)": "Faturamento Médio"})
            )
        
            # Formata visual
            df_resultado["Faturamento Médio"] = df_resultado["Faturamento Médio"].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )    
            ordem_dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]
            df_resultado["Dia da Semana"] = pd.Categorical(df_resultado["Dia da Semana"], categories=ordem_dias, ordered=True)
            df_resultado = df_resultado.sort_values(["Grupo", "Loja", "Dia da Semana"])
        
            # Exibe
            st.dataframe(df_resultado, use_container_width=True)
            import xlsxwriter
            from io import BytesIO
            
            # === Corrige valores para float antes de exportar
            df_export = df_resultado.copy()
            df_export["Faturamento Médio"] = (
                df_export["Faturamento Médio"]
                .astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            
            df_export["Faturamento Médio"] = pd.to_numeric(df_export["Faturamento Médio"], errors="coerce").fillna(0.0)
                        
            # Gera arquivo Excel com formatação
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet("Previsão FC")
            
            # Formatos
            formato_header = workbook.add_format({'bold': True, 'bg_color': '#CCCCCC'})
            formato_reais = workbook.add_format({'num_format': 'R$ #,##0.00'})
            
            # Escreve cabeçalhos
            for col_idx, col_name in enumerate(df_export.columns):
                worksheet.write(0, col_idx, col_name, formato_header)
            
            # Escreve dados com formatação de reais
            for row_idx, row in enumerate(df_export.itertuples(index=False), start=1):
                for col_idx, value in enumerate(row):
                    if df_export.columns[col_idx] == "Faturamento Médio":
                        worksheet.write_number(row_idx, col_idx, value, formato_reais)
                    else:
                        worksheet.write(row_idx, col_idx, value)
            
            # Ajusta largura
            for i, col in enumerate(df_export.columns):
                max_width = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, max_width)
            
            workbook.close()
            output.seek(0)
            
            # Botão de download
            st.download_button(
                label="⬇️ Baixar Excel",
                data=output,
                file_name="Previsao_FC.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )




        # === Conciliação Adquirente ===
        with aba_conciliacao:    
            st.warning("📌 em desenvolvimento")
    except Exception as e:
        st.error(f"❌ Erro ao acessar dados: {e}")


