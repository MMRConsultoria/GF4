import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import unicodedata
from io import BytesIO
import gspread
import os
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter, column_index_from_string
import psycopg2
from datetime import datetime, timedelta, date

st.set_page_config(page_title="Meio de Pagamento", layout="wide")

# ====== CSS para travar o botão 3S e estilizar abas (cole após imports) ======
st.markdown(
    """
    <style>
    /* Tabs styling */
    .stApp { background-color: #f9f9f9; }
    div[data-baseweb="tab-list"] { margin-top: 20px; }
    button[data-baseweb="tab"] {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 10px 20px;
        margin-right: 10px;
        transition: all 0.3s ease;
        font-size: 16px;
        font-weight: 600;
    }
    button[data-baseweb="tab"]:hover { background-color: #dce0ea; color: black; }
    button[data-baseweb="tab"][aria-selected="true"] { background-color: #0366d6; color: white; }

    /* Hide top toolbar */
    [data-testid="stToolbar"] { visibility: hidden; height: 0%; position: fixed; }
    .stSpinner { visibility: visible !important; }

    /* ===== Button container fixed (botao-vermelho) ===== */
    div.botao-vermelho {
        display: inline-flex !important;
        flex: 0 0 200px !important;    /* largura fixa: ajuste se quiser */
        min-width: 200px !important;
        max-width: 200px !important;
        box-sizing: border-box !important;
        margin: 6px 0 !important;
    }

    /* Botão ocupa 100% do container fixo e não quebra texto */
    div.botao-vermelho > button {
        width: 100% !important;
        height: 36px !important;
        min-height: 36px !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-sizing: border-box !important;

        /* Estética */
        background-color: #ff4b4b !important;
        color: white !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 0 10px !important;
    }

    div.botao-vermelho > button * {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }

    div.botao-vermelho > button:hover {
        background-color: #ff3333 !important;
        color: white !important;
    }

    /* ensure internal <p> etc inside streamlit button follow styles */
    div.botao-vermelho button p {
        margin: 0 !important;
        color: white !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# 🔒 Bloqueia o acesso caso o usuário não esteja logado
if not st.session_state.get("acesso_liberado"):
    st.stop()

# ✅ Inicializar controle de modo
if "modo_3s_mp" not in st.session_state:
    st.session_state.modo_3s_mp = False

# ======================
# Helpers de normalização
# ======================
def _strip_accents_keep_case(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s or "")).encode("ASCII", "ignore").decode("ASCII")

def _norm(s: str) -> str:
    s = _strip_accents_keep_case(s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def _is_formato2(df_headed: pd.DataFrame) -> bool:
    cols = {_norm(c) for c in df_headed.columns}
    return ("data" in cols and "total" in cols
            and any("cod" in c and "empresa" in c for c in cols)
            and any("forma" in c and "pag" in c for c in cols))

def _rename_cols_formato2(df: pd.DataFrame) -> pd.DataFrame:
    new_names = {}
    for c in df.columns:
        n = _norm(c)
        if "cod" in n and "empresa" in n:
            new_names[c] = "cod_empresa"
        elif n == "data":
            new_names[c] = "data"
        elif "forma" in n and "pag" in n:
            new_names[c] = "forma_pgto"
        elif "bandeira" in n:
            new_names[c] = "bandeira"
        elif "tipo" in n and "cart" in n:
            new_names[c] = "tipo_cartao"
        elif n == "total" or "valor" in n:
            new_names[c] = "total"
    return df.rename(columns=new_names)

# ======================
# Helper para escolher primeiro DF válido
# ======================
def first_nonempty_df(*candidates):
    for x in candidates:
        if x is None:
            continue
        if isinstance(x, pd.DataFrame):
            if not x.empty:
                return x
        else:
            return x
    return None

# ======================
# Leitura robusta por assinatura (xls/xlsx/xlsm)
# ======================
def _sniff_excel_kind(uploaded_file) -> str:
    try:
        pos = uploaded_file.tell()
    except Exception:
        pos = None
    try:
        uploaded_file.seek(0)
        head = uploaded_file.read(8)
    finally:
        try:
            uploaded_file.seek(pos or 0)
        except Exception:
            pass

    if not isinstance(head, (bytes, bytearray)):
        return "unknown"
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK\x05\x06") or head.startswith(b"PK\x07\x08"):
        return "xlsx"
    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls"
    return "unknown"

def read_excel_smart(file_or_xls, sheet_name=0, header=0):
    if isinstance(file_or_xls, pd.ExcelFile):
        return pd.read_excel(file_or_xls, sheet_name=sheet_name, header=header)
    kind = _sniff_excel_kind(file_or_xls)
    try:
        file_or_xls.seek(0)
    except Exception:
        pass
    try:
        if kind == "xls":
            return pd.read_excel(file_or_xls, sheet_name=sheet_name, header=header, engine="xlrd")
        return pd.read_excel(file_or_xls, sheet_name=sheet_name, header=header, engine="openpyxl")
    except Exception as e1:
        try:
            file_or_xls.seek(0)
        except Exception:
            pass
        alt = "openpyxl" if kind == "xls" else "xlrd"
        try:
            return pd.read_excel(file_or_xls, sheet_name=sheet_name, header=header, engine=alt)
        except Exception as e2:
            raise RuntimeError(f"Falha lendo Excel (kind={kind}). Tentativas: {e1} | {e2}")

def excel_file_smart(uploaded_file):
    kind = _sniff_excel_kind(uploaded_file)
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    eng = "xlrd" if kind == "xls" else "openpyxl"
    try:
        return pd.ExcelFile(uploaded_file, engine=eng)
    except Exception as e1:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
        alt = "openpyxl" if eng == "xlrd" else "xlrd"
        try:
            return pd.ExcelFile(uploaded_file, engine=alt)
        except Exception as e2:
            raise RuntimeError(f"Falha abrindo ExcelFile (kind={kind}). Tentativas: {e1} | {e2}")

# The above excel_file_smart had a tricky engine selection in original; preserve resilient behavior:
def excel_file_smart(uploaded_file):
    kind = _sniff_excel_kind(uploaded_file)
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    eng = "xlrd" if kind == "xls" else "openpyxl"
    try:
        return pd.ExcelFile(uploaded_file, engine=eng)
    except Exception as e1:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
        alt = "openpyxl" if eng == "xlrd" else "xlrd"
        try:
            return pd.ExcelFile(uploaded_file, engine=alt)
        except Exception as e2:
            raise RuntimeError(f"Falha abrindo ExcelFile (kind={kind}). Tentativas: {e1} | {e2}")

# ======================
# Processamento Formato 2 (plano)
# ======================
def processar_formato2(
    df_src: pd.DataFrame,
    df_empresa: pd.DataFrame,
    df_meio_pgto_google_norm: pd.DataFrame,
    depara_ciss_lookup: dict = None,
) -> pd.DataFrame:
    import re

    df = _rename_cols_formato2(df_src.copy())

    req = {"cod_empresa", "data", "forma_pgto", "bandeira", "tipo_cartao", "total"}
    faltando = [c for c in req if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas obrigatórias ausentes no arquivo: {faltando}")

    df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
    df["total"] = pd.to_numeric(df["total"], errors="coerce").fillna(0.0)

    ban_raw = df["bandeira"].fillna("").astype(str).str.strip()
    tip_raw = df["tipo_cartao"].fillna("").astype(str).str.strip()

    ban_norm = ban_raw.map(_norm)
    tip_norm = tip_raw.map(_norm)

    def _tipo_from_text(tn: str) -> str:
        if not tn:
            return ""
        if "cred" in tn:
            return "CREDITO"
        if "deb" in tn:
            return "DEBITO"
        return ""

    tip_label = pd.Series([_tipo_from_text(t) for t in tip_norm], index=df.index)
    tip_label = tip_label.where(
        tip_label != "",
        pd.Series([_tipo_from_text(b) for b in ban_norm], index=df.index)
    )

    tip_key = pd.Series(
        ["cred" if v == "CREDITO" else ("deb" if v == "DEBITO" else "") for v in tip_label],
        index=df.index
    )
    contains_tipo = pd.Series(
        [(k != "" and k in b) for b, k in zip(ban_norm, tip_key)],
        index=df.index
    )

    meio_composto = np.where(
        (ban_raw != "") | (tip_label != ""),
        np.where(
            contains_tipo,
            ban_raw,
            (ban_raw + " " + tip_label).str.strip()
        ),
        ""
    )

    fallback = (
        df["forma_pgto"].astype(str).str.strip()
          .str.replace(r"^\d+\s*-\s*", "", regex=True)
    )
    meio_composto = pd.Series(meio_composto, index=df.index).where(
        lambda s: s != "", fallback
    )

    df["Meio de Pagamento"] = meio_composto.map(_strip_accents_keep_case).str.strip()
    df["Meio de Pagamento"] = df["Meio de Pagamento"].str.replace(
        r'(?i)\b(\w+)(\s+\1\b)+', r'\1', regex=True
    )

    tipo_pgto_map = dict(
        zip(
            df_meio_pgto_google_norm["__meio_norm__"],
            df_meio_pgto_google_norm["Tipo de Pagamento"].astype(str),
        )
    )
    tipo_dre_map = dict(
        zip(
            df_meio_pgto_google_norm["__meio_norm__"],
            df_meio_pgto_google_norm["Tipo DRE"].astype(str),
        )
    )
    df["__meio_norm__"] = df["Meio de Pagamento"].map(_norm)
    df["Tipo de Pagamento"] = df["__meio_norm__"].map(tipo_pgto_map).fillna("")
    df["Tipo DRE"] = df["__meio_norm__"].map(tipo_dre_map).fillna("")
    df.drop(columns=["__meio_norm__"], inplace=True, errors="ignore")

    emp = df_empresa.copy()
    emp["Código Everest"] = emp["Código Everest"].astype(str).str.strip()
    df["Código Everest"] = df["cod_empresa"].astype(str).str.strip()
    df = df.merge(
        emp[["Código Everest", "Loja", "Grupo", "Código Grupo Everest"]],
        on="Código Everest",
        how="left",
    )

    df["Sistema"] = "CISS"

    dias_semana = {
        "Monday": "segunda-feira",
        "Tuesday": "terça-feira",
        "Wednesday": "quarta-feira",
        "Thursday": "quinta-feira",
        "Friday": "sexta-feira",
        "Saturday": "sábado",
        "Sunday": "domingo",
    }
    df["Dia da Semana"] = df["data"].dt.day_name().map(dias_semana)
    df["Mês"] = df["data"].dt.month.map(
        {1:"jan",2:"fev",3:"mar",4:"abr",5:"mai",6:"jun",7:"jul",8:"ago",9:"set",10:"out",11:"nov",12:"dez"}
    )
    df["Ano"] = df["data"].dt.year
    df["Data"] = df["data"].dt.strftime("%d/%m/%Y")

    df.rename(columns={"total": "Valor (R$)"}, inplace=True)
    col_order = [
        "Data","Dia da Semana",
        "Meio de Pagamento","Tipo de Pagamento","Tipo DRE",
        "Loja","Código Everest","Grupo","Código Grupo Everest",
        "Sistema",
        "Valor (R$)","Mês","Ano"
    ]
    for c in col_order:
        if c not in df.columns:
            df[c] = ""

    df_final = df[col_order].copy()

    try:
        df_final.sort_values(by=["Data", "Loja"], inplace=True)
    except Exception:
        pass

    return df_final

# ======================
# 3SCheckout - Certificado AWS + Postgres
# ======================
CERT_PATH = "aws-us-east-2-bundle.pem"

if "cert_written" not in st.session_state:
    with open(CERT_PATH, "w", encoding="utf-8") as f:
        f.write(st.secrets["certs"]["aws_rds_us_east_2"])
    st.session_state["cert_written"] = True

def get_db_conn():
    return psycopg2.connect(
        host=st.secrets["db"]["host"],
        port=st.secrets["db"]["port"],
        dbname=st.secrets["db"]["database"],
        user=st.secrets["db"]["user"],
        password=st.secrets["db"]["password"],
        sslmode="verify-full",
        sslrootcert=CERT_PATH,
    )

def parse_props(x):
    if pd.isna(x): return {}
    try:
        if isinstance(x, str):
            return json.loads(x)
    except:
        try:
            import ast
            return ast.literal_eval(x)
        except:
            return {}
    return x if isinstance(x, dict) else {}

def buscar_meio_pagamento_3s_checkout(df_empresa: pd.DataFrame, df_meio_pgto_google: pd.DataFrame):
    """Busca dados do 3S Checkout direto do banco e processa para Meio de Pagamento (últimos 60 dias até ontem)"""
    conn = get_db_conn()
    try:
        # Ajuste para fuso horário de Brasília (UTC-3) e define "ontem" como limite máximo
        agora_brasil = datetime.utcnow() - timedelta(hours=3)
        ontem = (agora_brasil - timedelta(days=1)).date()
        data_inicio = ontem - timedelta(days=59)  # últimos 60 dias incluindo ontem

        # BASE (order_picture) p/ VOID_TYPE + store/date
        query_op = """
            SELECT
                order_picture_id,
                store_code,
                business_dt,
                custom_properties
            FROM public.order_picture
            WHERE business_dt >= %s
              AND business_dt <= %s
              AND store_code NOT IN ('0000', '0001', '9999')
              AND state_id = 5
        """
        # Passa os dois parâmetros: data_inicio e ontem
        df_op = pd.read_sql(query_op, conn, params=(data_inicio, ontem))

        if df_op.empty:
            return None, f"Nenhum dado encontrado no período {data_inicio.strftime('%d/%m/%Y')} até {ontem.strftime('%d/%m/%Y')}", 0

        df_op["business_dt"] = pd.to_datetime(df_op["business_dt"], errors="coerce")
        df_op["store_code"] = df_op["store_code"].astype(str).str.lstrip("0").str.strip()

        props = df_op["custom_properties"].apply(parse_props)
        df_op["VOID_TYPE"] = props.apply(lambda x: x.get("VOID_TYPE") if isinstance(x, dict) else None)
        df_op = df_op[df_op["VOID_TYPE"].isna() | (df_op["VOID_TYPE"] == "") | (df_op["VOID_TYPE"] == 0)].copy()
        if df_op.empty:
            return None, "Todos os registros foram filtrados (VOID)", 0

        # TENDER (order_picture_tender) com troco
        query_tender = """
            SELECT
                order_picture_id,
                tender_amount,
                change_amount,
                details
            FROM public.order_picture_tender
            WHERE order_picture_id = ANY(%s)
        """
        op_ids = df_op["order_picture_id"].dropna().astype(int).tolist()
        df_tender = pd.read_sql(query_tender, conn, params=(op_ids,))
        if df_tender.empty:
            return None, "Nenhum tender encontrado", 0

        df_tender["tender_amount"] = pd.to_numeric(df_tender["tender_amount"], errors="coerce").fillna(0)
        df_tender["change_amount"] = pd.to_numeric(df_tender["change_amount"], errors="coerce").fillna(0)

        tender_props = df_tender["details"].apply(parse_props)
        df_tender["Meio de Pagamento"] = tender_props.apply(
            lambda x: x.get("tenderDescr") if isinstance(x, dict) else None
        )
        df_tender["tip_amount"] = pd.to_numeric(
            tender_props.apply(lambda x: x.get("tipAmount", 0) if isinstance(x, dict) else 0),
            errors="coerce"
        ).fillna(0)

        # líquido de troco
        df_tender["valor_liquido"] = (df_tender["tender_amount"] - df_tender["change_amount"]).clip(lower=0)
        df_tender["Valor (R$)"] = df_tender["valor_liquido"] + df_tender["tip_amount"]

        # junta com a base filtrada (VOID ok)
        df_tender = df_tender.merge(
            df_op[["order_picture_id", "store_code", "business_dt"]],
            on="order_picture_id",
            how="inner"
        )

        dias_traducao = {
            "Monday": "segunda-feira", "Tuesday": "terça-feira", "Wednesday": "quarta-feira",
            "Thursday": "quinta-feira", "Friday": "sexta-feira", "Saturday": "sábado", "Sunday": "domingo"
        }

        df_tender["Data_dt"] = pd.to_datetime(df_tender["business_dt"], errors="coerce").dt.normalize()
        df_tender["Data"] = df_tender["Data_dt"].dt.strftime("%d/%m/%Y")
        df_tender["Dia da Semana"] = df_tender["Data_dt"].dt.day_name().map(dias_traducao)

        # tabela empresa por Código
        emp = df_empresa.copy()
        if "Código Everest" in emp.columns:
            emp["Código Everest"] = (
                emp["Código Everest"].astype(str).str.replace(r"\D", "", regex=True).str.lstrip("0").str.strip()
            )
        df_tender["Código Everest"] = df_tender["store_code"].astype(str).str.replace(r"\D", "", regex=True).str.lstrip("0").str.strip()

        df_tender = df_tender.merge(
            emp[["Código Everest", "Loja", "Grupo", "Código Grupo Everest"]],
            on="Código Everest",
            how="left"
        )
        df_tender["Loja"] = df_tender["Loja"].astype(str).str.strip().str.lower()

        # mês/ano/sistema
        meses = {1:"jan",2:"fev",3:"mar",4:"abr",5:"mai",6:"jun",7:"jul",8:"ago",9:"set",10:"out",11:"nov",12:"dez"}
        df_tender["Mês"] = df_tender["Data_dt"].dt.month.map(meses)
        df_tender["Ano"] = df_tender["Data_dt"].dt.year
        df_tender["Sistema"] = "3SCheckout"

        # Tipo Pgto / Tipo DRE
        df_meio_pgto_google = df_meio_pgto_google.copy()
        if "__meio_norm__" not in df_meio_pgto_google.columns:
            df_meio_pgto_google["__meio_norm__"] = df_meio_pgto_google["Meio de Pagamento"].map(_norm)

        pgto_map = dict(zip(df_meio_pgto_google["__meio_norm__"], df_meio_pgto_google["Tipo de Pagamento"].astype(str)))
        dre_map  = dict(zip(df_meio_pgto_google["__meio_norm__"], df_meio_pgto_google["Tipo DRE"].astype(str)))

        df_tender["__meio_norm__"] = df_tender["Meio de Pagamento"].astype(str).str.strip().map(_norm)
        df_tender["Tipo de Pagamento"] = df_tender["__meio_norm__"].map(pgto_map).fillna("")
        df_tender["Tipo DRE"] = df_tender["__meio_norm__"].map(dre_map).fillna("")
        df_tender.drop(columns=["__meio_norm__"], inplace=True, errors="ignore")

        # RESUMO (dia + loja + meio)
        resumo = df_tender.groupby(
            ["Data_dt", "Data", "Dia da Semana", "Meio de Pagamento", "Tipo de Pagamento", "Tipo DRE",
             "Loja", "Código Everest", "Grupo", "Código Grupo Everest", "Mês", "Ano", "Sistema"],
            dropna=False,
            as_index=False
        ).agg({"Valor (R$)": "sum"})

        resumo["Valor (R$)"] = pd.to_numeric(resumo["Valor (R$)"], errors="coerce").fillna(0).round(2)

        # remove coluna auxiliar
        resumo.drop(columns=["Data_dt"], inplace=True)

        # ordem padrão
        col_order = [
            "Data", "Dia da Semana", "Meio de Pagamento", "Tipo de Pagamento", "Tipo DRE",
            "Loja", "Código Everest", "Grupo", "Código Grupo Everest",
            "Valor (R$)", "Mês", "Ano", "Sistema"
        ]
        for c in col_order:
            if c not in resumo.columns:
                resumo[c] = ""
        resumo = resumo[col_order]

        total_registros = len(df_op)
        return resumo, None, total_registros

    except Exception as e:
        return None, str(e), 0
    finally:
        conn.close()

# ======================
# Spinner + cargas do Google
# ======================
with st.spinner("⏳ Processando..."):
    # 🔌 Conexão com Google Sheets
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
    gc = gspread.authorize(credentials)
    planilha = gc.open("Tabelas")

    # Tabela Empresa
    df_empresa = pd.DataFrame(planilha.worksheet("Tabela Empresa").get_all_records())

    # Tabela Meio Pagamento (RAW = com duplicatas)
    df_meio_pgto_raw = pd.DataFrame(planilha.worksheet("Tabela Meio Pagamento").get_all_records())
    df_meio_pgto_raw.columns = [str(c).strip() for c in df_meio_pgto_raw.columns]
    for col in ["Meio de Pagamento", "Tipo de Pagamento", "Tipo DRE"]:
        if col not in df_meio_pgto_raw.columns:
            df_meio_pgto_raw[col] = ""
        df_meio_pgto_raw[col] = df_meio_pgto_raw[col].astype(str).str.strip()

    # LOOKUP De→para CiSS
    depara_col = None
    for c in df_meio_pgto_raw.columns:
        if "ciss" in _norm(c):
            depara_col = c
            break

    depara_ciss_lookup = {}
    if depara_col:
        tmp = df_meio_pgto_raw.copy()
        tmp["_depara_ciss_val_"] = tmp[depara_col].astype(str).str.strip()
        tmp = tmp[tmp["_depara_ciss_val_"] != ""]
        if not tmp.empty:
            tmp["_depara_key_"] = tmp["_depara_ciss_val_"].map(_norm)
            tmp["_canon_norm_"] = tmp["Meio de Pagamento"].astype(str).str.strip().map(_norm)

            conflitos = tmp.groupby("_depara_key_")["_canon_norm_"].nunique()
            conflitos = conflitos[conflitos > 1]
            if not conflitos.empty:
                st.warning(
                    "⚠️ Existem chaves de 'De para CISS' apontando para mais de um canônico. "
                    "Usando o primeiro encontrado para cada chave conflitante."
                )

            tmp_sorted = tmp.drop_duplicates(subset=["_depara_key_"], keep="first")
            map_key = tmp_sorted["_depara_key_"].tolist()
            map_val = tmp_sorted["Meio de Pagamento"].astype(str).str.strip().tolist()
            depara_ciss_lookup = dict(zip(map_key, map_val))

    # MAPAS de classificação (deduplicamos por canônico)
    df_meio_pgto_google = df_meio_pgto_raw.copy()
    df_meio_pgto_google["__meio_norm__"] = df_meio_pgto_google["Meio de Pagamento"].map(_norm)
    df_meio_pgto_google = df_meio_pgto_google.drop_duplicates(subset=["__meio_norm__"], keep="first")

    # 🔥 Título
    st.markdown("""
        <div style='display: flex; align-items: center; gap: 10px;'>
            <img src='https://img.icons8.com/color/48/graph.png' width='40'/>
            <h1 style='display: inline; margin: 0; font-size: 2.4rem;'>Meio de Pagamento</h1>
        </div>
    """, unsafe_allow_html=True)

    # ========================
    # 🗂️ Abas
    # ========================
    tab1, tab2 = st.tabs(["📥 Upload e Processamento", "🔄 Atualizar Google Sheets"])

    # ======================
    # 📥 Aba 1
    # ======================
    with tab1:
        # ========== BOTÃO 3S CHECKOUT ==========
        st.markdown("### 🔄 Atualização Automática 3S Checkout")

        # colocamos container HTML fixo ao redor do botão (garante largura fixa)
        st.markdown('<div style="display:flex; align-items:flex-start; gap:12px;">', unsafe_allow_html=True)
        st.markdown('<div class="botao-vermelho">', unsafe_allow_html=True)
        if st.button("🔄 Atualizar 3S Checkout", key="btn_3s_mp"):
            st.session_state.modo_3s_mp = True
            st.session_state.df_meio_pagamento = None  # limpa upload manual

            with st.spinner("Buscando dados do banco..."):
                resumo_3s, erro_3s, total_registros = buscar_meio_pagamento_3s_checkout(df_empresa, df_meio_pgto_google)

                if erro_3s:
                    st.error(f"❌ Erro ao buscar dados: {erro_3s}")
                elif resumo_3s is not None and not resumo_3s.empty:
                    st.session_state.resumo_3s_mp = resumo_3s
                    st.session_state.total_registros_3s_mp = total_registros
                    st.rerun()
                else:
                    st.warning("⚠️ Nenhum dado encontrado para o período.")
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # ========== EXIBIR RESULTADO 3S ==========
        if st.session_state.modo_3s_mp and "resumo_3s_mp" in st.session_state:
            resumo_3s = st.session_state.resumo_3s_mp
            total_registros = st.session_state.total_registros_3s_mp

            st.success(f"✅ {total_registros} registros processados com sucesso!")

            # Verificar meios não localizados
            meios_norm_tabela = set(df_meio_pgto_google["__meio_norm__"])
            meios_nao_localizados = resumo_3s[
                ~resumo_3s["Meio de Pagamento"].astype(str).str.strip().map(_norm).isin(meios_norm_tabela)
            ]["Meio de Pagamento"].astype(str).unique()

            if len(meios_nao_localizados) > 0:
                meios_nao_localizados_str = "<br>".join(meios_nao_localizados)
                mensagem = f"""
                ⚠️ {len(meios_nao_localizados)} meio(s) de pagamento não localizado(s):<br>{meios_nao_localizados_str}
                <br>✏️ Atualize a tabela clicando 
                <a href='https://docs.google.com/spreadsheets/d/1AVacOZDQT8vT-E8CiD59IVREe3TpKwE_25wjsj--qTU' target='_blank'><strong>aqui</strong></a>.
                """
                st.markdown(mensagem, unsafe_allow_html=True)
            else:
                st.success("✅ Todos os meios de pagamento foram localizados!")

            # Mostrar resumo do período
            datas_validas = pd.to_datetime(resumo_3s["Data"], format="%d/%m/%Y", errors='coerce').dropna()
            if not datas_validas.empty:
                data_inicial = datas_validas.min().strftime("%d/%m/%Y")
                data_final_str = datas_validas.max().strftime("%d/%m/%Y")
                valor_total = resumo_3s["Valor (R$)"].sum()
                valor_total_formatado = f"R$ {valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"""
                        <div style='font-size:1.2rem;'>📅 Período processado<br>{data_inicial} até {data_final_str}</div>
                    """, unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                        <div style='font-size:1.2rem;'>💰 Valor total<br><span style='color:green;'>{valor_total_formatado}</span></div>
                    """, unsafe_allow_html=True)

            # Gera Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                resumo_3s.to_excel(writer, sheet_name='Meio de Pagamento', index=False)
            output.seek(0)

            st.download_button(
                label="📥 Baixar Excel 3S Checkout",
                data=output,
                file_name=f"meio_pagamento_3s_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_3s_excel"
            )

            # Botão para voltar ao upload manual
            st.markdown("---")
            if st.button("↩️ Voltar para Upload Manual", key="btn_voltar_upload_manual"):
                st.session_state.modo_3s_mp = False
                if "resumo_3s_mp" in st.session_state:
                    del st.session_state.resumo_3s_mp
                if "total_registros_3s_mp" in st.session_state:
                    del st.session_state.total_registros_3s_mp
                st.rerun()

        # ========== UPLOAD MANUAL (só aparece se não estiver no modo 3S) ==========
        if not st.session_state.modo_3s_mp:
            st.markdown("---")
            st.markdown("### 📁 Upload Manual")

            uploaded_file = st.file_uploader(
                "📁 Clique para selecionar ou arraste aqui o arquivo Excel",
                type=["xlsx", "xlsm", "xls"],
                key="file_uploader_meio_pagamento"
            )

            if uploaded_file:
                try:
                    # Detecta Formato 2 pelo header real
                    df_head = read_excel_smart(uploaded_file, sheet_name=0, header=0)
                    if _is_formato2(df_head):
                        # ➜ Formato 2 (plano)
                        df_meio_pagamento = processar_formato2(
                            df_head,
                            df_empresa,
                            df_meio_pgto_google,
                            depara_ciss_lookup=depara_ciss_lookup
                        )
                    else:
                        # ➜ Formato 1 (layout antigo)
                        uploaded_file.seek(0)
                        xls = excel_file_smart(uploaded_file)
                        abas_disponiveis = xls.sheet_names

                        aba_escolhida = abas_disponiveis[0] if len(abas_disponiveis) == 1 else st.selectbox(
                            "Escolha a aba para processar", abas_disponiveis, key="select_aba_upload")

                        df_raw = read_excel_smart(xls, sheet_name=aba_escolhida, header=None)
                        df_raw = df_raw[~df_raw.iloc[:, 1].astype(str).str.lower().str.contains("total|subtotal", na=False)]

                        if str(df_raw.iloc[0, 1]).strip().lower() != "faturamento diário por meio de pagamento":
                            st.error("❌ A célula B1 deve conter 'Faturamento diário por meio de pagamento'.")
                            st.stop()

                        linha_inicio_dados, blocos, col, loja_atual = 5, [], 3, None
                        while col < df_raw.shape[1]:
                            valor_linha4 = str(df_raw.iloc[3, col]).strip()
                            match = re.match(r"^\d+\s*-\s*(.+)$", valor_linha4)
                            if match:
                                loja_atual = match.group(1).strip().lower()

                            meio_pgto = str(df_raw.iloc[4, col]).strip()
                            if not loja_atual or not meio_pgto or meio_pgto.lower() in ["nan", ""]:
                                col += 1; continue

                            linha3 = str(df_raw.iloc[2, col]).strip().lower()
                            linha5 = meio_pgto.lower()
                            if any(p in t for t in [linha3, valor_linha4.lower(), linha5] for p in ["total", "serv/tx", "total real"]):
                                col += 1; continue

                            try:
                                df_temp = df_raw.iloc[linha_inicio_dados:, [2, col]].copy()
                                df_temp.columns = ["Data", "Valor (R$)"]
                                df_temp = df_temp[~df_temp["Data"].astype(str).str.lower().str.contains("total|subtotal")]
                                df_temp.insert(1, "Meio de Pagamento", meio_pgto.lower())
                                df_temp.insert(2, "Loja", loja_atual)
                                blocos.append(df_temp)
                            except Exception as e:
                                st.warning(f"⚠️ Erro ao processar coluna {col}: {e}")
                            col += 1

                        if not blocos:
                            st.error("❌ Nenhum dado válido encontrado.")
                        else:
                            df_meio_pagamento = pd.concat(blocos, ignore_index=True).dropna()
                            df_meio_pagamento = df_meio_pagamento[~df_meio_pagamento["Data"].astype(str).str.lower().str.contains("total|subtotal")]
                            df_meio_pagamento["Data"] = pd.to_datetime(df_meio_pagamento["Data"], dayfirst=True, errors="coerce")
                            df_meio_pagamento = df_meio_pagamento[df_meio_pagamento["Data"].notna()]

                            # Derivados + join por Loja
                            dias_semana = {'Monday':'segunda-feira','Tuesday':'terça-feira','Wednesday':'quarta-feira',
                                           'Thursday':'quinta-feira','Friday':'sexta-feira','Saturday':'sábado','Sunday':'domingo'}
                            df_meio_pagamento["Dia da Semana"] = df_meio_pagamento["Data"].dt.day_name().map(dias_semana)
                            df_meio_pagamento = df_meio_pagamento.sort_values(by=["Data", "Loja"])
                            df_meio_pagamento["Data"] = df_meio_pagamento["Data"].dt.strftime("%d/%m/%Y")

                            df_meio_pagamento["Loja"] = df_meio_pagamento["Loja"].str.strip().str.replace(r"^\d+\s*-\s*", "", regex=True).str.lower()
                            df_empresa["Loja"] = df_empresa["Loja"].str.strip().str.lower()
                            df_meio_pagamento = pd.merge(df_meio_pagamento, df_empresa, on="Loja", how="left")

                            # Mapeia Tipo de Pagamento / Tipo DRE
                            if "Meio de Pagamento" not in df_meio_pagamento.columns:
                                df_meio_pagamento["Meio de Pagamento"] = ""
                            df_meio_pagamento["__meio_norm__"] = df_meio_pagamento["Meio de Pagamento"].map(_norm)
                            col_meio_idx = df_meio_pagamento.columns.get_loc("Meio de Pagamento")
                            df_meio_pagamento.insert(
                                loc=col_meio_idx + 1,
                                column="Tipo de Pagamento",
                                value=df_meio_pagamento["__meio_norm__"].map(
                                    dict(zip(df_meio_pgto_google["__meio_norm__"], df_meio_pgto_google["Tipo de Pagamento"]))
                                ).fillna("")
                            )
                            df_meio_pagamento.insert(
                                loc=col_meio_idx + 2,
                                column="Tipo DRE",
                                value=df_meio_pagamento["__meio_norm__"].map(
                                    dict(zip(df_meio_pgto_google["__meio_norm__"], df_meio_pgto_google["Tipo DRE"]))
                                ).fillna("")
                            )
                            df_meio_pagamento.drop(columns=["__meio_norm__"], inplace=True, errors="ignore")

                            df_meio_pagamento["Mês"] = pd.to_datetime(df_meio_pagamento["Data"], dayfirst=True).dt.month.map({
                                1:'jan',2:'fev',3:'mar',4:'abr',5:'mai',6:'jun',7:'jul',8:'ago',9:'set',10:'out',11:'nov',12:'dez'})
                            df_meio_pagamento["Ano"] = pd.to_datetime(df_meio_pagamento["Data"], dayfirst=True).dt.year

                            # ➕ Sistema (Formato 1)
                            df_meio_pagamento["Sistema"] = "Colibri"

                            # 💡 Ordem padrão de saída
                            col_order = [
                                "Data", "Dia da Semana",
                                "Meio de Pagamento", "Tipo de Pagamento", "Tipo DRE",
                                "Loja", "Código Everest", "Grupo", "Código Grupo Everest",
                                "Sistema",
                                "Valor (R$)", "Mês", "Ano"
                            ]
                            for c in ["Código Everest", "Grupo", "Código Grupo Everest"]:
                                if c not in df_meio_pagamento.columns:
                                    df_meio_pagamento[c] = ""
                            for c in col_order:
                                if c not in df_meio_pagamento.columns:
                                    df_meio_pagamento[c] = ""

                    # 🔁 Consolida duplicatas por Data + Loja + Meio de Pagamento
                    if 'df_meio_pagamento' in locals() and not df_meio_pagamento.empty:
                        tmp = df_meio_pagamento.copy()

                        if "Valor (R$)" in tmp.columns:
                            tmp["Valor (R$)"] = pd.to_numeric(tmp["Valor (R$)"], errors="coerce").fillna(0)

                        tmp["_k_data"] = pd.to_datetime(tmp["Data"], dayfirst=True, errors="coerce").dt.date
                        tmp["_k_loja"] = tmp["Loja"].astype(str).str.strip().str.lower()
                        tmp["_k_meio"] = tmp["Meio de Pagamento"].astype(str).str.strip().str.lower()

                        agg_dict = {
                            "Data": "first",
                            "Dia da Semana": "first",
                            "Meio de Pagamento": "first",
                            "Tipo de Pagamento": "first",
                            "Tipo DRE": "first",
                            "Loja": "first",
                            "Código Everest": "first",
                            "Grupo": "first",
                            "Código Grupo Everest": "first",
                            "Sistema": "first",
                            "Mês": "first",
                            "Ano": "first",
                            "Valor (R$)": "sum",
                        }

                        df_meio_pagamento = (
                            tmp.groupby(["_k_data", "_k_loja", "_k_meio"], as_index=False)
                               .agg(agg_dict)
                               .drop(columns=["_k_data", "_k_loja", "_k_meio"])
                        )

                        df_meio_pagamento = df_meio_pagamento.reindex(columns=[c for c in tmp.columns if c in df_meio_pagamento.columns])

                        if 'col_order' in locals():
                            df_meio_pagamento = df_meio_pagamento.reindex(columns=[c for c in col_order if c in df_meio_pagamento.columns])

                    # Resultado pronto
                    st.session_state.df_meio_pagamento = df_meio_pagamento

                    # KPIs topo
                    dts = pd.to_datetime(df_meio_pagamento["Data"], dayfirst=True, errors="coerce")
                    periodo_min = dts.min().strftime("%d/%m/%Y") if not dts.empty else ""
                    periodo_max = dts.max().strftime("%d/%m/%Y") if not dts.empty else ""
                    col1, col2 = st.columns(2)
                    col1.markdown(f"<div style='font-size:1.2rem;'>📅 Período processado<br>{periodo_min} até {periodo_max}</div>", unsafe_allow_html=True)
                    valor_total = f"R$ {df_meio_pagamento['Valor (R$)'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    col2.markdown(f"<div style='font-size:1.2rem;'>💰 Valor total<br><span style='color:green;'>{valor_total}</span></div>", unsafe_allow_html=True)

                    # Validações
                    empresas_nao_localizadas = df_meio_pagamento[
                        df_meio_pagamento["Loja"].astype(str).str.strip().isin(["", "nan"])
                    ]["Código Everest"].unique() if "Código Everest" in df_meio_pagamento.columns else []
                    meios_norm_tabela = set(df_meio_pgto_google["__meio_norm__"])
                    meios_nao_localizados = df_meio_pagamento[
                        ~df_meio_pagamento["Meio de Pagamento"].astype(str).str.strip().map(_norm).isin(meios_norm_tabela)
                    ]["Meio de Pagamento"].astype(str).unique()

                    # Exportar Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_meio_pagamento.to_excel(writer, index=False, sheet_name="FaturamentoPorMeio")
                        wb = writer.book
                        ws = wb["FaturamentoPorMeio"]

                        headers = [cell.value for cell in ws[1]]
                        if "Sistema" in headers:
                            sys_idx = headers.index("Sistema") + 1
                            target_idx = column_index_from_string("O")

                            col_vals = [ws.cell(row=r, column=sys_idx).value for r in range(1, ws.max_row + 1)]
                            ws.delete_cols(sys_idx)

                            if sys_idx < target_idx:
                                target_idx -= 1

                            for r, val in enumerate(col_vals, start=1):
                                ws.cell(row=r, column=target_idx, value=val)

                    output.seek(0)

                    st.download_button(
                        "📥 Baixar relatório Excel",
                        data=output,
                        file_name="FaturamentoPorMeio.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_relatorio_excel"
                    )

                    # Mensagens pós-validação
                    if len(empresas_nao_localizadas) == 0 and len(meios_nao_localizados) == 0:
                        st.success("✅ Todas as empresas e todos os meios de pagamento foram localizados!")
                    else:
                        if len(empresas_nao_localizadas) > 0:
                            empresas_nao_localizadas_str = "<br>".join(empresas_nao_localizadas)
                            st.markdown(f"""
                            ⚠️ {len(empresas_nao_localizadas)} Código(s) Everest sem correspondência:<br>{empresas_nao_localizadas_str}
                            <br>✏️ Atualize a tabela clicando 
                            <a href='https://docs.google.com/spreadsheets/d/1AVacOZDQT8vT-E8CiD59IVREe3TpKwE_25wjsj--qTU' target='_blank'><strong>aqui</strong></a>.
                            """, unsafe_allow_html=True)
                        if len(meios_nao_localizados) > 0:
                            meios_nao_localizados_str = "<br>".join(meios_nao_localizados)
                            st.markdown(f"""
                            ⚠️ {len(meios_nao_localizados)} meio(s) de pagamento não localizado(s):<br>{meios_nao_localizados_str}
                            <br>✏️ Atualize a tabela clicando 
                            <a href='https://docs.google.com/spreadsheets/d/1AVacOZDQT8vT-E8CiD59IVREe3TpKwE_25wjsj--qTU' target='_blank'><strong>aqui</strong></a>.
                            """, unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"❌ Erro ao processar: {e}")

    # ======================
    # 🔄 Aba 2
    # ======================
    with tab2:
        st.markdown("🔗 [Abrir planilha Faturamento Meio Pagamento](https://docs.google.com/spreadsheets/d/1GSI291SEeeU9MtOWkGwsKGCGMi_xXMSiQnL_9GhXxfU/edit?gid=1278257122#gid)")

        # ✅ Aceita TANTO df_meio_pagamento (upload) QUANTO resumo_3s_mp (3S Checkout)
        df_para_enviar = first_nonempty_df(
            st.session_state.get("resumo_3s_mp"),
            st.session_state.get("df_meio_pagamento")
        )

        if df_para_enviar is None or df_para_enviar.empty:
            st.warning("⚠️ Primeiro faça o upload e o processamento na Aba 1 ou rode a atualização 3S Checkout.")
        elif not all(c in df_para_enviar.columns for c in ["Meio de Pagamento","Loja","Data"]):
            st.warning("⚠️ O arquivo processado não tem as colunas necessárias.")
        else:
            df_final = df_para_enviar.copy()

            # Garantir "Tipo de Pagamento" e "Tipo DRE"
            df_final["Meio de Pagamento"] = (
                df_final["Meio de Pagamento"].astype(str).str.strip().str.lower()
            )

            df_meio_pgto_google.columns = [str(c).strip() for c in df_meio_pgto_google.columns]
            for c in ["Meio de Pagamento", "Tipo de Pagamento", "Tipo DRE"]:
                if c not in df_meio_pgto_google.columns:
                    df_meio_pgto_google[c] = ""

            df_meio_pgto_google["Meio de Pagamento"] = (
                df_meio_pgto_google["Meio de Pagamento"].astype(str).str.strip().str.lower()
            )
            df_meio_pgto_google["Tipo de Pagamento"] = (
                df_meio_pgto_google["Tipo de Pagamento"].astype(str).str.strip()
            )
            df_meio_pgto_google["Tipo DRE"] = (
                df_meio_pgto_google["Tipo DRE"].astype(str).str.strip()
            )

            pgto_map = dict(zip(df_meio_pgto_google["Meio de Pagamento"], df_meio_pgto_google["Tipo de Pagamento"]))
            dre_map  = dict(zip(df_meio_pgto_google["Meio de Pagamento"], df_meio_pgto_google["Tipo DRE"]))

            df_final["__meio_norm__"] = df_final["Meio de Pagamento"].astype(str).str.strip().str.lower()

            if "Tipo de Pagamento" not in df_final.columns:
                pos = df_final.columns.get_loc("Meio de Pagamento") + 1
                df_final.insert(pos, "Tipo de Pagamento", df_final["__meio_norm__"].map(pgto_map))
            else:
                df_final["Tipo de Pagamento"] = df_final["Tipo de Pagamento"].fillna(df_final["__meio_norm__"].map(pgto_map))

            if "Tipo DRE" not in df_final.columns:
                pos = df_final.columns.get_loc("Tipo de Pagamento") + 1
                df_final.insert(pos, "Tipo DRE", df_final["__meio_norm__"].map(dre_map))
            else:
                df_final["Tipo DRE"] = df_final["Tipo DRE"].fillna(df_final["__meio_norm__"].map(dre_map))

            df_final.drop(columns=["__meio_norm__"], inplace=True, errors="ignore")

            # Construir chave de duplicidade "M" (inclui Sistema)
            df_final['M'] = (
                pd.to_datetime(df_final['Data'], format='%d/%m/%Y').dt.strftime('%Y-%m-%d')
                + df_final['Meio de Pagamento'] + df_final['Loja'] + df_final['Sistema']
            )

            # valor para float
            df_final['Valor (R$)'] = df_final['Valor (R$)'].apply(lambda x: float(str(x).replace(',', '.')))
            # data para serial do Sheets
            df_final['Data'] = (pd.to_datetime(df_final['Data'], dayfirst=True) - pd.Timestamp("1899-12-30")).dt.days
            # inteiros
            for col in ["Código Everest", "Código Grupo Everest", "Ano"]:
                if col in df_final.columns:
                    df_final[col] = df_final[col].apply(lambda x: int(x) if pd.notnull(x) and str(x).strip() != "" else "")

            # Planilha destino
            sh_fatur = gc.open("Faturamento Meio Pagamento")
            aba_destino = sh_fatur.worksheet("Faturamento Meio Pagamento")
            #aba_destino = gc.open("Faturamento Meio Pagamento").worksheet("Faturamento Meio Pagamento")
            valores_existentes = aba_destino.get_all_values()

            if valores_existentes:
                header = [h.strip() for h in valores_existentes[0]]
            else:
                header = [
                    "Data","Dia da Semana","Meio de Pagamento","Tipo de Pagamento","Tipo DRE",
                    "Loja","Código Everest","Grupo","Código Grupo Everest",
                    "Valor (R$)","Mês","Ano","M","Sistema"
                ]

            for c in header:
                if c not in df_final.columns:
                    df_final[c] = ""

            df_final = df_final.reindex(columns=header, fill_value="")

            header_lower = [h.lower() for h in header]
            m_idx = header_lower.index("m") if "m" in header_lower else -1

            if len(valores_existentes) > 1:
                if m_idx >= 0:
                    dados_existentes = set(
                        linha[m_idx] for linha in valores_existentes[1:] if len(linha) > m_idx
                    )
                else:
                    dados_existentes = set(
                        linha[-1] for linha in valores_existentes[1:] if len(linha) > 0
                    )
            else:
                dados_existentes = set()

            novos_dados, duplicados = [], []
            for linha in df_final.fillna("").values.tolist():
                chave_m = linha[m_idx] if m_idx >= 0 else linha[-1]
                if chave_m not in dados_existentes:
                    novos_dados.append(linha)
                    dados_existentes.add(chave_m)
                else:
                    duplicados.append(linha)

            lojas_nao_cadastradas = df_final[df_final["Código Everest"].astype(str).isin(["", "nan"])]['Loja'].unique() \
                if "Código Everest" in df_final.columns else []
            todas_lojas_ok = len(lojas_nao_cadastradas) == 0

            
            # ====================================================== 
            # =========================================================
            # 📥 BOTÃO ÚNICO: ENVIAR E ATUALIZAR CACHE
            # =========================================================
            if st.button("📥 Enviar dados e Atualizar Cache", key="btn_enviar_e_cache"):
                with st.spinner("🔄 Processando envio e atualizando cache..."):
                    # 1. Envio dos novos dados
                    if novos_dados:
                        aba_destino.append_rows(novos_dados)
                        st.success(f"✅ {len(novos_dados)} novos registros enviados!")
                    else:
                        st.info("ℹ️ Nenhum novo registro para enviar.")
                    
                    if duplicados:
                        st.warning(f"⚠️ {len(duplicados)} registros duplicados ignorados.")

                    # 2. Rotina de Cache (Executa logo após o envio)
                    try:
                        # Recarrega os valores da aba principal (agora com os novos dados inclusos)
                        valores_origem = aba_destino.get_all_values()
                        header_cache = valores_origem[0]
                        dados_corpo = valores_origem[1:]
                        
                        df_temp = pd.DataFrame(dados_corpo, columns=header_cache)
                        
                        hoje = date.today()
                        primeiro_dia_mes_atual = date(hoje.year, hoje.month, 1)
                        ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
                        primeiro_dia_mes_anterior = date(ultimo_dia_mes_anterior.year, ultimo_dia_mes_anterior.month, 1)
                        
                        def converter_data_filtro(x):
                            try:
                                if str(x).isdigit(): return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(x))).date()
                                return pd.to_datetime(x, dayfirst=True).date()
                            except: return None

                        df_temp['__dt_filtro__'] = df_temp['Data'].apply(converter_data_filtro)
                        mask = (df_temp['__dt_filtro__'] >= primeiro_dia_mes_anterior) & (df_temp['__dt_filtro__'] <= ultimo_dia_mes_anterior)
                        dados_filtrados = df_temp.loc[mask].drop(columns=['__dt_filtro__']).values.tolist()

                        try:
                            sh_cache = sh_fatur.worksheet("CACHE_FILTRADO")
                        except:
                            sh_cache = sh_fatur.add_worksheet(title="CACHE_FILTRADO", rows="1000", cols=str(len(header_cache)))

                        sh_cache.clear()
                        lista_final = [header_cache] + dados_filtrados
                        
                        if len(lista_final) > 1:
                            sh_cache.update("A1", lista_final, value_input_option="USER_ENTERED")
                            st.info(f"✅ CACHE_FILTRADO atualizado com {len(dados_filtrados)} linhas do mês anterior.")
                        else:
                            sh_cache.update("A1", [header_cache])
                            st.warning("⚠️ Nenhuma linha do mês anterior encontrada para o cache.")

                    except Exception as e_cache:
                        st.error(f"⚠️ Dados enviados, mas falha ao atualizar cache: {e_cache}")
