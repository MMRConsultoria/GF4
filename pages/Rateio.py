# pages/PainelResultados.py
import streamlit as st
st.set_page_config(page_title="Vendas Diarias", layout="wide")

# ===== Imports =====
import pandas as pd
from io import BytesIO
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImg
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import pytz
import io

# 🔒 Bloqueio de acesso
if not st.session_state.get("acesso_liberado"):
    st.stop()

# ===== CSS (visual) =====
st.markdown("""
    <style>
        [data-testid="stToolbar"] { visibility: hidden; height: 0%; position: fixed; }
        .stSpinner { visibility: visible !important; }
        .stApp { background-color: #f9f9f9; }
        div[data-baseweb="tab-list"] { margin-top: 20px; }
        button[data-baseweb="tab"] {
            background-color: #f0f2f6; border-radius: 10px;
            padding: 10px 20px; margin-right: 10px;
            transition: all 0.3s ease; font-size: 16px; font-weight: 600;
        }
        button[data-baseweb="tab"]:hover { background-color: #dce0ea; color: black; }
        button[data-baseweb="tab"][aria-selected="true"] { background-color: #0366d6; color: white; }

        /* multiselect sem tags coloridas */
        div[data-testid="stMultiSelect"] [data-baseweb="tag"] { background-color: transparent !important; border: none !important; color: black !important; }
        div[data-testid="stMultiSelect"] [data-baseweb="tag"] * { color: black !important; fill: black !important; }
        div[data-testid="stMultiSelect"] > div { background-color: transparent !important; }
    </style>
""", unsafe_allow_html=True)

# ===== Cabeçalho =====
st.markdown("""
    <div style='display: flex; align-items: center; gap: 10px; margin-bottom: 12px;'>
        <img src='https://img.icons8.com/color/48/graph.png' width='40'/>
        <h1 style='display: inline; margin: 0; font-size: 2.0rem;'>Rateio</h1>
    </div>
""", unsafe_allow_html=True)

with st.spinner("⏳ Processando..."):
    # ===== Conexão + dados base (comum às abas) =====
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
    gc = gspread.authorize(credentials)
    planilha = gc.open("Vendas diarias")

    df_empresa = pd.DataFrame(planilha.worksheet("Tabela Empresa").get_all_records())
    df_vendas  = pd.DataFrame(planilha.worksheet("Fat Sistema Externo").get_all_records())

    # Normalização comum
    df_empresa.columns = df_empresa.columns.str.strip()
    df_vendas.columns  = df_vendas.columns.str.strip()
    if "Loja" in df_empresa.columns:
        df_empresa["Loja"] = df_empresa["Loja"].astype(str).str.strip().str.upper()
    if "Grupo" in df_empresa.columns:
        df_empresa["Grupo"] = df_empresa["Grupo"].astype(str).str.strip()
    if "Data" in df_vendas.columns:
        df_vendas["Data"] = pd.to_datetime(df_vendas["Data"], dayfirst=True, errors="coerce")
    if "Loja" in df_vendas.columns:
        df_vendas["Loja"] = df_vendas["Loja"].astype(str).str.strip().str.upper()
    if "Grupo" in df_vendas.columns:
        df_vendas["Grupo"] = df_vendas["Grupo"].astype(str).str.strip()
    # Merge com Tipo
    if {"Loja","Tipo"}.issubset(df_empresa.columns):
        df_vendas = df_vendas.merge(df_empresa[["Loja","Tipo"]], on="Loja", how="left")
    else:
        df_vendas["Tipo"] = df_vendas.get("Tipo", "")

    # Converter Fat.Total para número
    if "Fat.Total" in df_vendas.columns:
        df_vendas["Fat.Total"] = (
            df_vendas["Fat.Total"].astype(str)
            .str.replace("R$", "", regex=False)
            .str.replace("(", "-", regex=False)
            .str.replace(")", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_vendas["Fat.Total"] = pd.to_numeric(df_vendas["Fat.Total"], errors="coerce").fillna(0.0)
    else:
        df_vendas["Fat.Total"] = 0.0

    # ===== Abas =====
    aba1, aba2 = st.tabs(["📄 %Faturamento", "🔄 Volumetria"])

    # ----------------------------------------------------------------------
    # ABA 1 - % FATURAMENTO  (tudo desta aba fica aqui dentro)
    # ----------------------------------------------------------------------
    with aba1:
        # --------- Filtros ---------
        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            tipos = sorted(df_vendas["Tipo"].dropna().unique())
            tipos.insert(0, "Todos")
            tipo_sel = st.selectbox("🏪 Tipo:", options=tipos, index=0, key="tipo_fat")

        with col2:
            grupos = sorted(df_vendas["Grupo"].dropna().unique())
            grupos.insert(0, "Todos")
            grupo_sel = st.selectbox("👥 Grupo:", options=grupos, index=0, key="grupo_fat")

        with col3:
            df_vendas["Mes/Ano"] = df_vendas["Data"].dt.strftime("%m/%Y")

            def _ord_key(mmyyyy: str):
                try:
                    return datetime.strptime("01/" + str(mmyyyy), "%d/%m/%Y")
                except Exception:
                    return datetime.min

            meses_opts = sorted([m for m in df_vendas["Mes/Ano"].dropna().unique()], key=_ord_key)
            mes_atual = datetime.today().strftime("%m/%Y")
            default_meses = [mes_atual] if meses_opts and mes_atual in meses_opts else (meses_opts[-1:] if meses_opts else [])

            if meses_opts:
                meses_sel = st.multiselect(
                    "🗓️ Selecione os meses:",
                    options=meses_opts,
                    default=default_meses,
                    key="ms_meses_fat"
                )
            else:
                st.warning("⚠️ Nenhum mês disponível nos dados (verifique a coluna 'Data').")
                meses_sel = []

        # --------- Aplica filtros ---------
        if meses_sel:
            df_f = df_vendas[df_vendas["Mes/Ano"].isin(meses_sel)].copy()
        else:
            df_f = df_vendas.iloc[0:0].copy()

        df_f["Período"] = df_f["Data"].dt.strftime("%m/%Y")
        if tipo_sel != "Todos":
            df_f = df_f[df_f["Tipo"] == tipo_sel]
        if grupo_sel != "Todos":
            df_f = df_f[df_f["Grupo"] == grupo_sel]

        # --------- Agrupamento ---------
        metric = "Fat.Total"  # métrica desta aba
        if grupo_sel == "Todos":
            chaves = ["Tipo", "Grupo"]
        else:
            chaves = ["Grupo", "Loja"]

        df_ag = df_f.groupby(chaves + ["Período"], as_index=False)[metric].sum()
        df_fin = df_ag.groupby(chaves, as_index=False)[metric].sum().rename(columns={metric: "Total"})
        df_fin["Rateio"] = 0.0

        # --------- % e Subtotais ---------
        if grupo_sel == "Todos":
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            subt = df_fin.groupby("Tipo")["Total"].sum().reset_index().sort_values(by="Total", ascending=False)
            ordem_tipos = subt["Tipo"].tolist()
            df_fin["ord_tipo"] = df_fin["Tipo"].apply(lambda x: ordem_tipos.index(x) if x in ordem_tipos else 999)
            df_fin = df_fin.sort_values(by=["ord_tipo", "Total"], ascending=[True, False]).drop(columns="ord_tipo")

            linhas = []
            for t in ordem_tipos:
                bloco = df_fin[df_fin["Tipo"] == t].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Tipo", "Grupo"]).sum(numeric_only=True)
                subtotal["Tipo"] = t
                subtotal["Grupo"] = f"Subtotal {t}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)
        else:
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            df_fin = df_fin.sort_values(by=["Grupo", "Total"], ascending=[True, False])

            linhas = []
            for g in df_fin["Grupo"].unique():
                bloco = df_fin[df_fin["Grupo"] == g].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Grupo", "Loja"]).sum(numeric_only=True)
                subtotal["Grupo"] = g
                subtotal["Loja"] = f"Subtotal {g}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)

        # --------- TOTAL no topo ---------
        cols_drop = [c for c in ["Tipo","Grupo","Loja"] if c in df_fin.columns]
        apenas = df_fin.copy()
        for c in cols_drop:
            apenas = apenas[~apenas[c].astype(str).str.startswith("Subtotal", na=False)]
        linha_total = apenas.drop(columns=cols_drop, errors="ignore").sum(numeric_only=True)
        for c in cols_drop:
            linha_total[c] = ""
        linha_total[cols_drop[0] if cols_drop else "Grupo"] = "TOTAL"
        df_fin = pd.concat([pd.DataFrame([linha_total]), df_fin], ignore_index=True)

        # --------- RATEIO ---------
        df_fin["% Total"] = 0.0
        df_fin["Rateio"] = 0.0

        if grupo_sel == "Todos":
            def moeda_para_float(s: str) -> float:
                try: return float(s.replace(".", "").replace(",", "."))
                except: return 0.0

            tipos_unicos = [t for t in df_fin["Tipo"].dropna().unique()
                            if str(t).strip() not in ["", "TOTAL"] and not str(t).startswith("Subtotal")]
            valores_rateio = {}
            COLS_POR_LINHA = 3
            for i in range(0, len(tipos_unicos), COLS_POR_LINHA):
                linha = tipos_unicos[i:i+COLS_POR_LINHA]
                cols = st.columns(len(linha))
                for c, t in zip(cols, linha):
                    with c:
                        valor_str = st.text_input(f"💰 Rateio — {t}", value="0,00", key=f"rateio_{t}_fat")
                        valores_rateio[t] = moeda_para_float(valor_str)

            for t in df_fin["Tipo"].unique():
                mask = ((df_fin["Tipo"] == t) &
                        (~df_fin["Grupo"].astype(str).str.startswith("Subtotal")) &
                        (df_fin["Grupo"] != "TOTAL"))
                subtotal_t = df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Total"].sum()
                if subtotal_t > 0:
                    df_fin.loc[mask, "% Total"] = (df_fin.loc[mask, "Total"] / subtotal_t) * 100
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "% Total"] = 100

                valor_rateio = valores_rateio.get(t, 0.0)
                df_fin.loc[mask, "Rateio"] = df_fin.loc[mask, "% Total"] / 100 * valor_rateio
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Rateio"] = df_fin.loc[mask, "Rateio"].sum()
        else:
            total_rateio = st.number_input(f"💰 Rateio — {grupo_sel}",
                                           min_value=0.0, step=100.0, format="%.2f",
                                           key=f"rateio_{grupo_sel}_fat")
            mask_lojas = ((df_fin["Grupo"] == grupo_sel) &
                          (~df_fin["Loja"].astype(str).str.startswith("Subtotal")) &
                          (df_fin["Loja"] != "TOTAL"))
            subtotal_g = df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Total"].sum()
            if subtotal_g > 0:
                df_fin.loc[mask_lojas, "% Total"] = (df_fin.loc[mask_lojas, "Total"] / subtotal_g) * 100
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "% Total"] = 100
                df_fin.loc[mask_lojas, "Rateio"] = df_fin.loc[mask_lojas, "% Total"] / 100 * total_rateio
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Rateio"] = df_fin.loc[mask_lojas, "Rateio"].sum()
        # === Reordenar colunas (Aba 1) ===
        if grupo_sel == "Todos":
            col_order = ["Tipo", "Grupo", "Total", "% Total", "Rateio"]
        else:
            # quando filtra um grupo específico, aparece "Loja"
            col_order = ["Grupo", "Loja", "Total", "% Total", "Rateio"]
        
        # mantém só as colunas nessa ordem (as que existirem)
        df_fin = df_fin[[c for c in col_order if c in df_fin.columns]]
        # --------- Visual ---------
        df_view = df_fin.copy()
        def fmt_moeda(v):
            try: return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except: return v
        for c in ["Total","Rateio"]:
            if c in df_view.columns:
                df_view[c] = df_view[c].apply(lambda x: fmt_moeda(x) if pd.notnull(x) and x != "" else x)
        if "% Total" in df_view.columns:
            df_view["% Total"] = pd.to_numeric(df_view["% Total"], errors="coerce").apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")

        def aplicar_estilo_fat(df_in):
            def estilo(row):
                if "Grupo" in df_in.columns and row["Grupo"] == "TOTAL":
                    return ["background-color: #f4b084; font-weight: bold"] * len(row)
                if "Loja" in df_in.columns and isinstance(row.get("Loja",""), str) and row["Loja"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                if "Grupo" in df_in.columns and isinstance(row.get("Grupo",""), str) and row["Grupo"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                return ["" for _ in row]
            return df_in.style.apply(estilo, axis=1)

        st.dataframe(aplicar_estilo_fat(df_view), use_container_width=True, height=700)

        # --------- Exportar Excel ---------
        df_excel = df_fin.copy()
        if "% Total" in df_excel.columns:
            df_excel["% Total"] = pd.to_numeric(df_excel["% Total"], errors="coerce") / 100
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name="Relatório")
        out.seek(0)
        wb = load_workbook(out); ws = wb["Relatório"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center_alignment; cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            grupo_val = None
            try: grupo_val = row[1].value
            except: pass
            estilo_fundo = None
            if isinstance(grupo_val, str):
                if grupo_val.strip().upper() == "TOTAL": estilo_fundo = PatternFill("solid", fgColor="F4B084")
                elif "SUBTOTAL" in grupo_val.strip().upper(): estilo_fundo = PatternFill("solid", fgColor="D9D9D9")
            for cell in row:
                cell.border = border; cell.alignment = center_alignment
                if estilo_fundo: cell.fill = estilo_fundo
                col_name = ws.cell(row=1, column=cell.column).value
                if isinstance(cell.value, (int,float)):
                    cell.number_format = '0.00%' if col_name == "% Total" else '"R$" #,##0.00'

        for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        for col_nome in ["Tipo","Grupo","Loja"]:
            if col_nome in df_excel.columns:
                col_idx = df_excel.columns.get_loc(col_nome) + 1
                for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for c in cell: c.alignment = Alignment(horizontal="left")

        out_final = BytesIO(); wb.save(out_final); out_final.seek(0)
        st.download_button("📥 Baixar Excel", data=out_final,
                           file_name="Resumo_%Faturamento.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_fat")

        # --------- Exportar PDF ---------
        usuario = st.session_state.get("usuario_logado", "Usuário Desconhecido")
        sele = meses_sel
        if not sele: mes_rateio = "(sem dados)"
        elif len(sele) == 1: mes_rateio = sele[0]
        elif len(sele) == 2: mes_rateio = f"{sele[0]} e {sele[1]}"
        else: mes_rateio = f"{', '.join(sele[:-1])} e {sele[-1]}"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=20, rightMargin=20)
        elems = []; estilos = getSampleStyleSheet(); normal = estilos["Normal"]; h1 = estilos["Heading1"]
        try:
            logo_url = "https://raw.githubusercontent.com/MMRConsultoria/mmr-site/main/logo_grupofit.png"
            img = RLImg(logo_url, width=100, height=40); elems.append(img)
        except: pass
        elems.append(Paragraph(f"<b>Rateio - {mes_rateio}</b>", h1))
        fuso = pytz.timezone("America/Sao_Paulo")
        data_ger = datetime.now(fuso).strftime("%d/%m/%Y %H:%M")
        elems.append(Paragraph(f"<b>Usuário:</b> {usuario}", normal))
        elems.append(Paragraph(f"<b>Data de Geração:</b> {data_ger}", normal)); elems.append(Spacer(1,12))
        dados = [df_view.columns.tolist()] + df_view.values.tolist()
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#003366")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(1,1),(-1,-1),"CENTER"),
            ("ALIGN",(0,0),(0,-1),"LEFT"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,0),8),
            ("GRID",(0,0),(-1,-1),0.25,colors.grey),
        ]))
        for i in range(1, len(dados)):
            txt = str(dados[i][1]).strip().lower() if len(dados[i])>1 else ""
            if "subtotal" in txt or txt == "total":
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#BFBFBF")),
                                            ("FONTNAME",(0,i),(-1,i),"Helvetica-Bold")]))
            else:
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F2F2F2"))]))
        elems.append(tabela); doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        st.download_button("📄 Baixar PDF", data=pdf_bytes,
                           file_name=f"Rateio_%Faturamento_{datetime.now().strftime('%Y%m%d')}.pdf",
                           mime="application/pdf", key="dl_pdf_fat")


    # ----------------------------------------------------------------------
    # ABA 2 - VOLUMETRIA (Mês, Ano, Grupo, Funcionarios)
    # ----------------------------------------------------------------------
    with aba2:
        # ========= 1) Ler planilha externa com cabeçalhos (Mês, Ano, Grupo, Funcionarios) =========
        url_vol = "https://docs.google.com/spreadsheets/d/1AVacOZDQT8vT-E8CiD59IVREe3TpKwE_25wjsj--qTU/edit?gid=1461552258#gid=1461552258"
        gid_vol = 1461552258
    
        ss_vol = gc.open_by_url(url_vol)
        try:
            ws_vol = ss_vol.get_worksheet_by_id(gid_vol)
        except Exception:
            ws_vol = ss_vol.get_worksheet(0)
    
        rows = ws_vol.get_all_values()
        if not rows:
            st.error("Planilha de Volumetria vazia.")
            st.stop()
    
        headers = rows[0]
        data_rows = rows[1:]
    
        import unicodedata
        def norm(s):
            s = str(s or "")
            s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
            return s.strip().lower()
    
        # mapeia índices pelos nomes (tolerante a acentos/maiúsculas)
        idx = {norm(h): i for i, h in enumerate(headers)}
        def col_idx(*names):
            for n in names:
                n_norm = norm(n)
                if n_norm in idx:
                    return idx[n_norm]
            return None
    
        i_mes  = col_idx("mes", "mês")
        i_ano  = col_idx("ano")
        i_grp  = col_idx("grupo")
        i_func = col_idx("funcionarios", "funcionários", "qtd funcionarios", "qtde funcionarios")
    
        if None in (i_mes, i_ano, i_grp, i_func):
            st.error("Não encontrei uma ou mais colunas: Mês, Ano, Grupo, Funcionarios na planilha de Volumetria.")
            st.stop()
    
        def parse_mes_num(x):
            s = str(x or "").strip()
            if s.isdigit():
                m = int(s)
                return m if 1 <= m <= 12 else None
            # tenta nomes pt (JAN, JANEIRO...)
            s2 = norm(s).upper()
            mapa = {
                "JAN":1, "JANEIRO":1, "FEV":2, "FEVEREIRO":2, "MAR":3, "MARCO":3, "MARÇO":3,
                "ABR":4, "ABRIL":4, "MAI":5, "MAIO":5, "JUN":6, "JUNHO":6, "JUL":7, "JULHO":7,
                "AGO":8, "AGOSTO":8, "SET":9, "SETEMBRO":9, "OUT":10, "OUTUBRO":10,
                "NOV":11, "NOVEMBRO":11, "DEZ":12, "DEZEMBRO":12
            }
            return mapa.get(s2, None)
    
        def to_float(x):
            s = str(x or "").strip().replace(".", "").replace(",", ".")
            try:
                return float(s)
            except:
                return 0.0
    
        reg = []
        for r in data_rows:
            mes = parse_mes_num(r[i_mes] if len(r) > i_mes else "")
            ano = str(r[i_ano] if len(r) > i_ano else "").strip()
            grp = str(r[i_grp] if len(r) > i_grp else "").strip().upper()
            fun = to_float(r[i_func] if len(r) > i_func else "")
            if mes and ano and grp:
                reg.append({"Mes/Ano": f"{mes:02d}/{ano}", "Grupo": grp, "Funcionarios": fun})
    
        df_funcs = pd.DataFrame(reg)
        # agrega funcionários por mês+grupo
        if not df_funcs.empty:
            df_funcs = df_funcs.groupby(["Mes/Ano","Grupo"], as_index=False)["Funcionarios"].sum()
    
        # ========= 2) Filtros (meses = união faturamento + volumetria) =========
        # ========= 2) Filtros (meses = união faturamento + volumetria) =========
        # cria as colunas de filtros SÓ desta aba
        col1, col2, col3 = st.columns([1, 1, 2])
        
        # Mes/Ano nas duas bases
        df_vendas["Mes/Ano"] = df_vendas["Data"].dt.strftime("%m/%Y")
        
        with col1:
            # UNIÃO dos tipos do faturamento + o tipo "ADM"
            tipos_fat = set(df_vendas["Tipo"].dropna().astype(str).str.strip())
            tipos = sorted(tipos_fat | {"ADM"})
            tipos.insert(0, "Todos")
            tipo_sel = st.selectbox("🏪 Tipo:", options=tipos, index=0, key="tipo_vol")
        
        with col2:
            # UNIÃO de grupos: faturamento + volumetria
            grupos_fat = set(df_vendas["Grupo"].dropna().astype(str).str.strip().str.upper())
            grupos_vol = set(df_funcs["Grupo"].dropna().astype(str).str.strip().str.upper()) if not df_funcs.empty else set()
            grupos = sorted(grupos_fat | grupos_vol)
            grupos.insert(0, "Todos")
            grupo_sel = st.selectbox("👥 Grupo:", options=grupos, index=0, key="grupo_vol")
        
        with col3:
            from datetime import datetime
            def _ordkey(mmyyyy: str):
                try: return datetime.strptime("01/" + str(mmyyyy), "%d/%m/%Y")
                except Exception: return datetime.min
        
            meses_fat = set(df_vendas["Mes/Ano"].dropna().unique().tolist())
            meses_fun = set(df_funcs["Mes/Ano"].dropna().unique().tolist()) if not df_funcs.empty else set()
            meses_opts = sorted(meses_fat | meses_fun, key=_ordkey)
        
            mes_atual = datetime.today().strftime("%m/%Y")
            default_meses = [mes_atual] if mes_atual in meses_opts else (meses_opts[-1:] if meses_opts else [])
            meses_sel = st.multiselect(
                "🗓️ Selecione os meses:",
                options=meses_opts,
                default=default_meses,
                key="ms_meses_vol"
            )

    
        # ========= 3) Base de faturamento (para manter grupos do mês e exibir o valor) =========
        if meses_sel:
            df_f = df_vendas[df_vendas["Mes/Ano"].isin(meses_sel)].copy()
        else:
            df_f = df_vendas.iloc[0:0].copy()
    
        if tipo_sel != "Todos":
            df_f = df_f[df_f["Tipo"] == tipo_sel]
        if grupo_sel != "Todos":
            df_f = df_f[df_f["Grupo"].str.upper() == grupo_sel]
    
        df_f["Grupo"] = df_f["Grupo"].astype(str).str.strip().str.upper()
        df_fat = (
            df_f.groupby(["Tipo","Grupo"], as_index=False)["Fat.Total"]
                .sum().rename(columns={"Fat.Total":"Faturamento"})
        )
    
        # ========= 4) Funcionários por Grupo (somando meses selecionados) =========
        if not df_funcs.empty and meses_sel:
            df_funcs_sel = df_funcs[df_funcs["Mes/Ano"].isin(meses_sel)].copy()
        else:
            df_funcs_sel = df_funcs.iloc[0:0].copy() if not df_funcs.empty else pd.DataFrame(columns=["Grupo","Funcionarios"])
    
        df_fun_g = df_funcs_sel.groupby("Grupo", as_index=False)["Funcionarios"].sum() if not df_funcs_sel.empty \
                   else pd.DataFrame({"Grupo": [], "Funcionarios": []})
    
        # ========= 5) Juntar (outer): inclui grupos que existem só na volumetria -> Tipo = ADM =========
        df_fin = df_fat.merge(df_fun_g, on="Grupo", how="outer")
        df_fin["Tipo"] = df_fin["Tipo"].fillna("ADM")
        df_fin["Faturamento"] = df_fin["Faturamento"].fillna(0.0)
        df_fin["Funcionarios"] = df_fin["Funcionarios"].fillna(0.0)
    
        # aplica filtros depois da união
        if tipo_sel != "Todos":
            df_fin = df_fin[df_fin["Tipo"] == tipo_sel]
        if grupo_sel != "Todos":
            df_fin = df_fin[df_fin["Grupo"] == grupo_sel]
    
        # ========= 6) % por funcionários (dentro de cada Tipo) =========
        df_fin["% Total"] = 0.0
        for t in df_fin["Tipo"].dropna().unique():
            mask_t = (df_fin["Tipo"] == t)
            total_fun_t = df_fin.loc[mask_t, "Funcionarios"].sum()
            if total_fun_t > 0:
                df_fin.loc[mask_t, "% Total"] = df_fin.loc[mask_t, "Funcionarios"] / total_fun_t * 100
    
        # ========= 7) Subtotais por Tipo (ordena por Funcionários) =========
        ordem_tipos = (df_fin.groupby("Tipo")["Funcionarios"].sum()
                       .sort_values(ascending=False).index.tolist())
        df_fin["ord_tipo"] = df_fin["Tipo"].apply(lambda x: ordem_tipos.index(x) if x in ordem_tipos else 999)
        df_fin = df_fin.sort_values(["ord_tipo","Funcionarios"], ascending=[True, False]).drop(columns="ord_tipo")
    
        linhas = []
        for t in ordem_tipos:
            bloco = df_fin[df_fin["Tipo"] == t].copy()
            if bloco.empty: 
                continue
            linhas.append(bloco)
            sub = bloco.drop(columns=["Tipo","Grupo"]).sum(numeric_only=True)
            sub["Tipo"] = t
            sub["Grupo"] = f"Subtotal {t}"
            linhas.append(pd.DataFrame([sub]))
        if linhas:
            df_fin = pd.concat(linhas, ignore_index=True)
    
        # ========= 8) Linha TOTAL no topo =========
        cols_drop = [c for c in ["Tipo","Grupo"] if c in df_fin.columns]
        apenas = df_fin.copy()
        for c in cols_drop:
            apenas = apenas[~apenas[c].astype(str).str.startswith("Subtotal", na=False)]
        linha_total = apenas.drop(columns=cols_drop, errors="ignore").sum(numeric_only=True)
        for c in cols_drop:
            linha_total[c] = ""
        linha_total[cols_drop[0] if cols_drop else "Grupo"] = "TOTAL"
        df_fin = pd.concat([pd.DataFrame([linha_total]), df_fin], ignore_index=True)
    
        # ========= 9) INPUT MANUAL: Total a ratear (R$) =========
        # ========= 9) INPUT MANUAL: Despesa + Total a ratear (R$) =========
        c1, c2 = st.columns([2, 1])
        
        with c1:
            despesa_str = st.text_input(
                "🧾 Despesa",
                value=st.session_state.get("despesa_vol_str", ""),
                key="despesa_vol_str"
            )
        
        with c2:
            valor_str = st.text_input(
                "📦 Total a ratear (R$)",
                value="0,00",
                key="rateio_total_vol_str"
            )
        
        def moeda_ptbr_to_float(s: str) -> float:
            s = str(s or "").strip()
            s = s.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return float(s)
            except:
                return 0.0
        
        total_rateio = moeda_ptbr_to_float(valor_str)
        
        # >>> Rateio per capita: (total_rateio ÷ total_funcionarios) × funcionarios_do_grupo
        mask_regular = (
            ~df_fin["Grupo"].astype(str).str.startswith("Subtotal")
            & (df_fin["Grupo"].astype(str) != "TOTAL")
            & (df_fin["Tipo"].astype(str)  != "TOTAL")
        )
        
        # garante numérico
        df_fin["Funcionarios"] = pd.to_numeric(df_fin["Funcionarios"], errors="coerce").fillna(0)
        
        total_func = df_fin.loc[mask_regular, "Funcionarios"].sum()
        valor_por_func = (total_rateio / total_func) if total_func > 0 else 0.0
        
        # aplica o per-capita nos grupos
        df_fin.loc[mask_regular, "Rateio"] = df_fin.loc[mask_regular, "Funcionarios"] * valor_por_func
        
        # Subtotais por Tipo (somatório do rateio das linhas do tipo)
        for t in df_fin["Tipo"].dropna().unique():
            mask_tipo_regular   = mask_regular & (df_fin["Tipo"] == t)
            mask_subtotal_tipo  = (df_fin["Grupo"].astype(str) == f"Subtotal {t}")
            df_fin.loc[mask_subtotal_tipo, "Rateio"] = df_fin.loc[mask_tipo_regular, "Rateio"].sum()
        
        # Linha TOTAL
        mask_total_grupo = (df_fin["Grupo"].astype(str) == "TOTAL")
        mask_total_tipo  = (df_fin["Tipo"].astype(str)  == "TOTAL")
        soma_rateio = df_fin.loc[mask_regular, "Rateio"].sum()
        if mask_total_grupo.any():
            df_fin.loc[mask_total_grupo, "Rateio"] = soma_rateio
        elif mask_total_tipo.any():
            df_fin.loc[mask_total_tipo, "Rateio"] = soma_rateio
        
        # arredonda para 2 casas (centavos)
        df_fin["Rateio"] = df_fin["Rateio"].round(2)
        
        # Nova coluna com a descrição da despesa
        df_fin["Despesa"] = (despesa_str or "").strip()

        
        # ========= 10) Visão: esconder %/Faturamento; renomear Funcionários (com acento) =========
        # ========= 10) Visão: esconder %/Faturamento; renomear Funcionários; incluir Despesa =========
        df_view = df_fin.rename(columns={"Funcionarios": "Funcionários"})[
            ["Tipo", "Grupo", "Funcionários", "Despesa", "Rateio"]
        ].copy()
       
    
        # --------- Visual (Funcionários em inteiro; Rateio em R$) ---------
        def fmt_moeda_br(v):
            try:
                return "R$ " + f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                return v
        
        def fmt_int_br(v):
            try:
                return f"{int(round(float(v))):,}".replace(",", ".")
            except:
                return v
        
        if "Funcionários" in df_view.columns:
            df_view["Funcionários"] = df_view["Funcionários"].apply(lambda x: fmt_int_br(x) if pd.notnull(x) and x != "" else x)
        if "Rateio" in df_view.columns:
            df_view["Rateio"] = df_view["Rateio"].apply(lambda x: fmt_moeda_br(x) if pd.notnull(x) and x != "" else x)
        
        def aplicar_estilo_vol(df_in):
            def estilo(row):
                if row.get("Grupo","") == "TOTAL":
                    return ["background-color: #f4b084; font-weight: bold"] * len(row)
                if isinstance(row.get("Grupo",""), str) and row["Grupo"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                return ["" for _ in row]
            return df_in.style.apply(estilo, axis=1)
        
        st.dataframe(aplicar_estilo_vol(df_view), use_container_width=True, height=700)

    
        # --------- Exportar Excel (Tipo, Grupo, Funcionários, Rateio) ---------
        # --------- Exportar Excel (Aba 2: Tipo, Grupo, Funcionários, Despesa, Rateio) ---------
        df_excel = df_fin.rename(columns={"Funcionarios": "Funcionários"})[
            ["Tipo", "Grupo", "Funcionários", "Despesa", "Rateio"]
        ].copy()
        
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name="Relatório")
        out.seek(0)
        wb = load_workbook(out); ws = wb["Relatório"]
        
        # (deixe a parte de formatação igual: cabeçalho azul, bordas, auto-width etc.)
        # Só garanta que a formatação numérica fique assim:
        # - Funcionários: '#,##0'
        # - Rateio: '"R$" #,##0.00'

        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center_alignment; cell.border = border
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            try:
                grupo_val = row[1].value  # col B = Grupo
            except:
                grupo_val = None
            estilo_fundo = None
            if isinstance(grupo_val, str):
                if grupo_val.strip().upper() == "TOTAL":
                    estilo_fundo = PatternFill("solid", fgColor="F4B084")
                elif "SUBTOTAL" in grupo_val.strip().upper():
                    estilo_fundo = PatternFill("solid", fgColor="D9D9D9")
        
            for cell in row:
                cell.border = border; cell.alignment = center_alignment
                if estilo_fundo: cell.fill = estilo_fundo
                # formatos: Funcionários = inteiro; Rateio = moeda R$
                col_name = ws.cell(row=1, column=cell.column).value
                if isinstance(cell.value, (int,float)):
                    if col_name == "Rateio":
                        cell.number_format = '"R$" #,##0.00'
                    elif col_name == "Funcionários":
                        cell.number_format = '#,##0'
        
        # auto width
        for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2
        
        # alinha textos à esquerda
        for col_nome in ["Tipo","Grupo"]:
            if col_nome in df_excel.columns:
                col_idx = df_excel.columns.get_loc(col_nome) + 1
                for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for c in cell: c.alignment = Alignment(horizontal="left")
        
        out_final = BytesIO(); wb.save(out_final); out_final.seek(0)
        st.download_button("📥 Baixar Excel", data=out_final,
                           file_name="Resumo_Volumetria.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_vol")

    
        # ========= 13) Exportar PDF (usa df_view: já sem %/Faturamento) =========
        usuario = st.session_state.get("usuario_logado", "Usuário Desconhecido")
        sele = meses_sel
        if not sele: mes_lbl = "(sem dados)"
        elif len(sele) == 1: mes_lbl = sele[0]
        elif len(sele) == 2: mes_lbl = f"{sele[0]} e {sele[1]}"
        else: mes_lbl = f"{', '.join(sele[:-1])} e {sele[-1]}"
    
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=20, rightMargin=20)
        elems = []; estilos = getSampleStyleSheet(); normal = estilos["Normal"]; h1 = estilos["Heading1"]
        try:
            logo_url = "https://raw.githubusercontent.com/MMRConsultoria/mmr-site/main/logo_grupofit.png"
            img = RLImg(logo_url, width=100, height=40); elems.append(img)
        except: pass
        elems.append(Paragraph(f"<b>Volumetria por Funcionários - {mes_lbl}</b>", h1))
        fuso = pytz.timezone("America/Sao_Paulo")
        data_ger = datetime.now(fuso).strftime("%d/%m/%Y %H:%M")
        elems.append(Paragraph(f"<b>Usuário:</b> {usuario}", normal))
        elems.append(Paragraph(f"<b>Data de Geração:</b> {data_ger}", normal)); elems.append(Spacer(1,12))
    
        dados = [df_view.columns.tolist()] + df_view.values.tolist()
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#003366")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(1,1),(-1,-1),"CENTER"),
            ("ALIGN",(0,0),(0,-1),"LEFT"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,0),8),
            ("GRID",(0,0),(-1,-1),0.25,colors.grey),
        ]))
        for i in range(1, len(dados)):
            txt = str(dados[i][1]).strip().lower() if len(dados[i])>1 else ""
            if "subtotal" in txt or txt == "total":
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#BFBFBF")),
                                            ("FONTNAME",(0,i),(-1,i),"Helvetica-Bold")]))
            else:
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F2F2F2"))]))
        elems.append(tabela); doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        st.download_button("📄 Baixar PDF", data=pdf_bytes,
                            file_name=f"Volumetria_{datetime.now().strftime('%Y%m%d')}.pdf",
                            mime="application/pdf", key="dl_pdf_vol")
